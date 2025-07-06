"""
Research Gap Finder - Integrated Professional Academic Research Tool
Combines best features for paper search and gap analysis
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import sys
import os
from pathlib import Path
import json
import yaml
import time
import logging
import hashlib
import re
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
import io
import zipfile
import threading
from collections import defaultdict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import xlsxwriter
    EXCEL_AVAILABLE = True
except ImportError:
    logging.warning("Excel packages not installed - Excel export disabled")
    EXCEL_AVAILABLE = False

# Import secure configuration
try:
    from config import config, get_available_providers
    SECURE_CONFIG_AVAILABLE = True
    logging.info("Secure configuration loaded successfully")
except ImportError:
    SECURE_CONFIG_AVAILABLE = False
    logging.warning("Secure config not available - using environment variables directly")

# Try to import optional packages
try:
    from tenacity import retry, stop_after_attempt, wait_exponential
    TENACITY_AVAILABLE = True
except ImportError:
    logging.warning("Tenacity not installed - retry logic disabled")
    TENACITY_AVAILABLE = False
    # Create dummy decorator
    def retry(*args, **kwargs):
        def decorator(func):
            return func
        return decorator
    
    class stop_after_attempt:
        def __init__(self, n):
            pass
    
    class wait_exponential:
        def __init__(self, **kwargs):
            pass

# Configure logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Page configuration
st.set_page_config(
    page_title="Research Gap Finder",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="auto",  # Auto-collapse on mobile
    menu_items={
        'Get Help': 'https://github.com/calvinliao223/gap_finder_bot_GCIT',
        'Report a bug': 'https://github.com/calvinliao223/gap_finder_bot_GCIT/issues',
        'About': """
        # Research Gap Finder 🔬

        AI-powered tool for discovering research opportunities across multiple academic databases.

        **Features:**
        - Multi-source paper search (Semantic Scholar, Crossref)
        - Intelligent gap analysis with AI
        - Comprehensive paper grading system
        - Export capabilities (JSON, YAML, CSV, Excel)
        - Cross-platform responsive design

        **Optimized for:**
        - 📱 Mobile devices (iOS/Android)
        - 📟 Tablets (iPad/Android tablets)
        - 💻 Desktop computers (Windows/Mac/Linux)
        - 🌐 All major browsers (Chrome, Firefox, Safari, Edge)
        """
    }
)

# Professional CSS styling with responsive design
st.markdown("""
<style>
    /* ===== MOBILE-FIRST RESPONSIVE DESIGN ===== */

    /* Base styles for mobile devices */
    .main .block-container {
        padding-top: 1rem;
        padding-bottom: 1rem;
        padding-left: 1rem;
        padding-right: 1rem;
        max-width: 100%;
    }

    /* Header styling - Mobile first */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem 1rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }

    .main-header h1 {
        margin: 0;
        font-size: 1.8rem;
        font-weight: 700;
        line-height: 1.2;
    }

    .main-header p {
        margin: 0.5rem 0 0 0;
        font-size: 1rem;
        opacity: 0.95;
        line-height: 1.4;
    }

    /* ===== TABLET STYLES (768px and up) ===== */
    @media (min-width: 768px) {
        .main .block-container {
            padding-top: 1.5rem;
            padding-bottom: 1.5rem;
            padding-left: 2rem;
            padding-right: 2rem;
            max-width: 1000px;
        }

        .main-header {
            padding: 2rem;
            border-radius: 15px;
            margin-bottom: 2rem;
        }

        .main-header h1 {
            font-size: 2.2rem;
        }

        .main-header p {
            font-size: 1.1rem;
        }
    }

    /* ===== DESKTOP STYLES (1024px and up) ===== */
    @media (min-width: 1024px) {
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
            max-width: 1200px;
        }

        .main-header h1 {
            font-size: 2.5rem;
        }

        .main-header p {
            font-size: 1.2rem;
        }
    }
    
    /* ===== CHAT MESSAGE STYLING - RESPONSIVE ===== */

    /* Mobile-first chat messages */
    .chat-message {
        padding: 1rem;
        margin: 0.8rem 0;
        border-radius: 12px;
        border-left: 4px solid;
        background: white;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        transition: all 0.2s;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    .chat-message:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.12);
    }

    .user-message {
        border-left-color: #3b82f6;
        background: linear-gradient(to right, #f0f9ff, #ffffff);
        margin-left: 0.5rem;
    }

    .assistant-message {
        border-left-color: #059669;
        background: linear-gradient(to right, #f0fdf4, #ffffff);
        margin-right: 0.5rem;
    }

    /* Tablet styles for chat messages */
    @media (min-width: 768px) {
        .chat-message {
            padding: 1.2rem 1.5rem;
            margin: 1rem 0;
        }

        .user-message {
            margin-left: 1rem;
        }

        .assistant-message {
            margin-right: 1rem;
        }
    }

    /* Desktop styles for chat messages */
    @media (min-width: 1024px) {
        .user-message {
            margin-left: 2rem;
        }

        .assistant-message {
            margin-right: 2rem;
        }
    }
    
    /* ===== PAPER CARD STYLING - RESPONSIVE ===== */

    /* Mobile-first paper cards */
    .paper-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #e5e7eb;
        margin: 0.8rem 0;
        transition: all 0.2s;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    .paper-card:hover {
        border-color: #667eea;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.1);
    }

    .paper-title {
        font-weight: 600;
        color: #1e293b;
        font-size: 1rem;
        margin-bottom: 0.5rem;
        line-height: 1.4;
    }

    .paper-metadata {
        color: #64748b;
        font-size: 0.85rem;
        line-height: 1.6;
    }

    /* Tablet styles for paper cards */
    @media (min-width: 768px) {
        .paper-card {
            padding: 1.3rem;
        }

        .paper-title {
            font-size: 1.05rem;
        }

        .paper-metadata {
            font-size: 0.9rem;
        }
    }

    /* Desktop styles for paper cards */
    @media (min-width: 1024px) {
        .paper-card {
            padding: 1.5rem;
        }

        .paper-title {
            font-size: 1.1rem;
        }
    }
    
    /* ===== ENHANCED PAPER CARD - RESPONSIVE ===== */

    /* Mobile-first enhanced paper cards */
    .enhanced-paper-card {
        background: white;
        padding: 1rem;
        border-radius: 12px;
        border: 1px solid #e5e7eb;
        margin: 0.8rem 0;
        transition: all 0.3s;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    .enhanced-paper-card:hover {
        border-color: #667eea;
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.15);
        transform: translateY(-2px);
    }

    /* Tablet styles for enhanced paper cards */
    @media (min-width: 768px) {
        .enhanced-paper-card {
            padding: 1.3rem;
            margin: 1rem 0;
        }
    }

    /* Desktop styles for enhanced paper cards */
    @media (min-width: 1024px) {
        .enhanced-paper-card {
            padding: 1.5rem;
        }
    }

    /* Paper header - responsive */
    .paper-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 0.8rem;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid #f1f5f9;
        flex-wrap: wrap;
        gap: 0.5rem;
    }

    .paper-rank {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 700;
        white-space: nowrap;
    }

    .paper-grade {
        font-weight: 700;
        font-size: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        white-space: nowrap;
    }

    /* Tablet and desktop paper header */
    @media (min-width: 768px) {
        .paper-header {
            margin-bottom: 1rem;
            flex-wrap: nowrap;
        }

        .paper-rank {
            font-size: 0.85rem;
        }

        .paper-grade {
            font-size: 1.1rem;
        }
    }

    .quality-summary {
        background: #f8fafc;
        padding: 0.8rem;
        border-radius: 8px;
        margin-top: 1rem;
        font-style: italic;
        color: #475569;
        border-left: 3px solid #e2e8f0;
    }

    /* Graded Paper Card for Assessment Results */
    .graded-paper-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #e5e7eb;
        margin: 1rem 0;
        transition: all 0.3s;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    .grade-breakdown {
        background: #f8fafc;
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        font-size: 0.9rem;
    }

    .grade-explanation {
        color: #475569;
        font-style: italic;
        margin-top: 0.8rem;
        padding: 0.8rem;
        background: #f1f5f9;
        border-radius: 6px;
    }

    .strengths {
        color: #059669;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }

    /* ===== ENHANCED GAP ANALYSIS CARDS - RESPONSIVE ===== */

    /* Mobile-first enhanced gap cards */
    .enhanced-gap-card {
        background: linear-gradient(135deg, #fef3c7 0%, #fef9e7 100%);
        padding: 1.2rem;
        border-radius: 12px;
        border: 2px solid #fbbf24;
        margin: 1rem 0;
        position: relative;
        overflow: hidden;
        transition: all 0.3s;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    .enhanced-gap-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(251, 191, 36, 0.2);
    }

    /* Tablet styles for enhanced gap cards */
    @media (min-width: 768px) {
        .enhanced-gap-card {
            padding: 1.5rem;
            margin: 1.2rem 0;
            border-radius: 15px;
        }

        .enhanced-gap-card:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(251, 191, 36, 0.2);
        }
    }

    /* Desktop styles for enhanced gap cards */
    @media (min-width: 1024px) {
        .enhanced-gap-card {
            padding: 1.8rem;
        }
    }

    /* Gap header - responsive */
    .gap-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 0.8rem;
        padding-bottom: 0.6rem;
        border-bottom: 2px solid #fbbf24;
        flex-wrap: wrap;
        gap: 0.5rem;
    }

    .gap-type {
        background: #f59e0b;
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 700;
        display: inline-block;
        white-space: nowrap;
    }

    .gap-priority {
        font-size: 0.75rem;
        color: #92400e;
        font-weight: 600;
        background: rgba(255, 255, 255, 0.7);
        padding: 0.25rem 0.6rem;
        border-radius: 12px;
        white-space: nowrap;
    }

    /* Tablet and desktop gap header */
    @media (min-width: 768px) {
        .gap-header {
            margin-bottom: 1rem;
            padding-bottom: 0.8rem;
            flex-wrap: nowrap;
        }

        .gap-type {
            padding: 0.4rem 1rem;
            border-radius: 25px;
            font-size: 0.9rem;
        }

        .gap-priority {
            font-size: 0.85rem;
            padding: 0.3rem 0.8rem;
            border-radius: 15px;
        }
    }

    /* Legacy gap card for backward compatibility */
    .gap-card {
        background: linear-gradient(135deg, #fef3c7 0%, #fef9e7 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border: 2px solid #fbbf24;
        margin: 1rem 0;
        position: relative;
        overflow: hidden;
    }

    /* Show more sections */
    .show-more-section {
        background: linear-gradient(135deg, #e0f2fe 0%, #f0f9ff 100%);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        border-left: 4px solid #0ea5e9;
        margin: 1rem 0;
        text-align: center;
        transition: all 0.3s;
    }

    .show-more-section:hover {
        background: linear-gradient(135deg, #bae6fd 0%, #e0f2fe 100%);
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(14, 165, 233, 0.15);
    }

    .show-more-section p {
        margin: 0.3rem 0;
        color: #0c4a6e;
    }

    .show-more-section em {
        color: #0369a1;
        font-size: 0.9rem;
    }

    /* ===== TOUCH-FRIENDLY INPUT AND BUTTON STYLING ===== */

    /* Enhanced chat input styling - Mobile first */
    .stChatInput {
        border-radius: 12px;
        border: 2px solid #e5e7eb;
        transition: all 0.3s;
        min-height: 44px; /* iOS minimum touch target */
    }

    .stChatInput:focus-within {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
    }

    .stChatInput textarea {
        border: none !important;
        font-size: 16px !important; /* Prevents zoom on iOS */
        line-height: 1.5;
        min-height: 44px !important;
        padding: 12px !important;
    }

    .stChatInput button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        border: none !important;
        border-radius: 10px !important;
        transition: all 0.3s !important;
        min-width: 44px !important; /* iOS minimum touch target */
        min-height: 44px !important;
        padding: 12px !important;
    }

    .stChatInput button:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3) !important;
    }

    /* Button styling - Touch-friendly */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 12px 16px; /* Increased for better touch targets */
        font-weight: 600;
        transition: all 0.3s;
        text-transform: none;
        min-height: 44px; /* iOS minimum touch target */
        min-width: 44px;
        font-size: 14px;
        line-height: 1.4;
        cursor: pointer;
        -webkit-tap-highlight-color: rgba(102, 126, 234, 0.2);
    }

    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
    }

    .stButton > button:active {
        transform: translateY(0);
        box-shadow: 0 2px 8px rgba(102, 126, 234, 0.3);
    }

    /* Tablet and desktop button improvements */
    @media (min-width: 768px) {
        .stChatInput textarea {
            font-size: 1rem !important;
        }

        .stButton > button {
            padding: 0.8rem 1.4rem;
            font-size: 15px;
        }
    }

    @media (min-width: 1024px) {
        .stButton > button {
            padding: 0.9rem 1.6rem;
            font-size: 16px;
        }
    }
    
    /* Loading animation */
    .thinking-indicator {
        display: flex;
        align-items: center;
        gap: 0.8rem;
        padding: 1.2rem;
        background: linear-gradient(135deg, #f3f4f6 0%, #e5e7eb 100%);
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .thinking-dots {
        display: flex;
        gap: 0.3rem;
    }
    
    .thinking-dot {
        width: 10px;
        height: 10px;
        background: #667eea;
        border-radius: 50%;
        animation: thinking 1.4s infinite ease-in-out;
    }
    
    .thinking-dot:nth-child(1) { animation-delay: -0.32s; }
    .thinking-dot:nth-child(2) { animation-delay: -0.16s; }
    .thinking-dot:nth-child(3) { animation-delay: 0s; }
    
    @keyframes thinking {
        0%, 80%, 100% { 
            transform: scale(0.8); 
            opacity: 0.5; 
        }
        40% { 
            transform: scale(1.2); 
            opacity: 1; 
        }
    }
    
    /* Progress bar styling */
    .stProgress > div > div {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Provider status */
    .provider-badge {
        padding: 0.4rem 0.8rem;
        border-radius: 8px;
        font-size: 0.85rem;
        font-weight: 600;
        display: inline-block;
        margin: 0.2rem;
    }

    .provider-active {
        background: #10b981;
        color: white;
    }

    .provider-inactive {
        background: #ef4444;
        color: white;
    }

    /* ===== RESPONSIVE SIDEBAR STYLING ===== */

    /* Mobile sidebar adjustments */
    .css-1d391kg { /* Streamlit sidebar container */
        width: 100% !important;
        min-width: 100% !important;
    }

    /* Sidebar content spacing for mobile */
    .css-1d391kg .stMarkdown {
        margin-bottom: 0.5rem;
    }

    .css-1d391kg .stButton {
        margin-bottom: 0.5rem;
    }

    /* Sidebar columns for mobile - stack vertically */
    .css-1d391kg .row-widget {
        flex-direction: column;
    }

    .css-1d391kg .row-widget > div {
        width: 100% !important;
        margin-bottom: 0.5rem;
    }

    /* Tablet sidebar improvements */
    @media (min-width: 768px) {
        .css-1d391kg {
            width: 300px !important;
            min-width: 300px !important;
        }

        .css-1d391kg .row-widget {
            flex-direction: row;
        }

        .css-1d391kg .row-widget > div {
            width: 48% !important;
            margin-right: 4%;
            margin-bottom: 0.8rem;
        }

        .css-1d391kg .row-widget > div:last-child {
            margin-right: 0;
        }
    }

    /* Desktop sidebar */
    @media (min-width: 1024px) {
        .css-1d391kg {
            width: 320px !important;
            min-width: 320px !important;
        }

        .css-1d391kg .stMarkdown {
            margin-bottom: 1rem;
        }

        .css-1d391kg .stButton {
            margin-bottom: 0.8rem;
        }
    }

    /* ===== CROSS-BROWSER COMPATIBILITY FIXES ===== */

    /* Firefox specific fixes */
    @-moz-document url-prefix() {
        .stButton > button {
            -moz-appearance: none;
        }

        .stChatInput textarea {
            -moz-appearance: none;
        }
    }

    /* Safari specific fixes */
    @supports (-webkit-appearance: none) {
        .stButton > button {
            -webkit-appearance: none;
        }

        .stChatInput textarea {
            -webkit-appearance: none;
        }

        /* Fix for Safari touch events */
        .stButton > button {
            -webkit-touch-callout: none;
            -webkit-user-select: none;
        }
    }

    /* Edge/IE specific fixes */
    @supports (-ms-ime-align: auto) {
        .stButton > button {
            border: none;
        }
    }

    /* General cross-browser button fixes */
    .stButton > button {
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
        text-rendering: optimizeLegibility;
    }

    /* Ensure consistent box-sizing across browsers */
    *, *::before, *::after {
        box-sizing: border-box;
    }

    /* Fix for Windows high DPI displays */
    @media (-webkit-min-device-pixel-ratio: 1.25), (min-resolution: 120dpi) {
        .main-header h1 {
            font-weight: 600; /* Slightly lighter for high DPI */
        }

        .stButton > button {
            font-weight: 500;
        }
    }
</style>
""", unsafe_allow_html=True)

# Constants
SEMANTIC_SCHOLAR_API = "https://api.semanticscholar.org/graph/v1/paper/search"
CROSSREF_API = "https://api.crossref.org/works"
MAX_PAPERS_PER_SOURCE = 20
CACHE_DURATION = 3600  # 1 hour

# Rate limiting constants
SEMANTIC_SCHOLAR_RATE_LIMIT = 1.1  # seconds between requests (1 req/sec + buffer)
MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 2

# Enums and Data Classes
class ConversationPhase(Enum):
    """Tracks the current phase of research conversation"""
    INITIALIZATION = "initialization"
    TOPIC_EXPLORATION = "topic_exploration"
    DOMAIN_REFINEMENT = "domain_refinement"
    LITERATURE_REVIEW = "literature_review"
    PAPER_ANALYSIS = "paper_analysis"
    GAP_ANALYSIS = "gap_analysis"
    RECOMMENDATION_REVIEW = "recommendation_review"

class UserExpertiseLevel(Enum):
    """User's research expertise level"""
    BEGINNER = "beginner"
    INTERMEDIATE = "intermediate"
    ADVANCED = "advanced"
    EXPERT = "expert"

class PaperGrade(Enum):
    """Paper quality grades"""
    A_PLUS = "A+"
    A = "A"
    A_MINUS = "A-"
    B_PLUS = "B+"
    B = "B"
    B_MINUS = "B-"
    C_PLUS = "C+"
    C = "C"
    C_MINUS = "C-"
    D = "D"
    F = "F"

@dataclass
class Paper:
    """Represents an academic paper"""
    title: str
    authors: List[str]
    year: int
    venue: str
    abstract: str
    citations: int
    doi: Optional[str] = None
    url: Optional[str] = None
    relevance_score: float = 0.0
    source: str = ""
    paper_id: str = field(default_factory=lambda: "")
    
    def __post_init__(self):
        if not self.paper_id:
            self.paper_id = hashlib.md5(f"{self.title}{self.year}".encode()).hexdigest()

@dataclass
class PaperQualityScore:
    """Comprehensive paper quality assessment"""
    overall_grade: PaperGrade
    overall_score: float  # 0-10 scale
    methodology_score: float  # 0-10
    citation_impact_score: float  # 0-10
    novelty_score: float  # 0-10
    clarity_score: float  # 0-10
    applicability_score: float  # 0-10
    explanation: str
    strengths: List[str] = field(default_factory=list)
    weaknesses: List[str] = field(default_factory=list)

@dataclass
class ResearchGap:
    """Represents an identified research gap with enhanced analysis"""
    gap_type: str
    description: str
    supporting_evidence: List[str]
    potential_impact: str
    suggested_approach: str
    confidence_score: float
    novelty_grade: str = "Medium"  # High/Medium/Low
    impact_grade: str = "Medium"  # High/Medium/Low
    research_suggestions: List[str] = field(default_factory=list)
    methodology_suggestions: List[str] = field(default_factory=list)
    required_expertise: UserExpertiseLevel = UserExpertiseLevel.INTERMEDIATE
    estimated_timeline: str = "6-12 months"
    potential_collaborators: List[str] = field(default_factory=list)
    required_resources: List[str] = field(default_factory=list)

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    logger.warning("python-dotenv not installed, using system environment variables")

# Core Classes
class LLMProviderManager:
    """Manages multiple LLM providers with fallback"""
    
    def __init__(self):
        self.providers = {}
        self.current_provider = None
        self._initialize_providers()
        
        # Apply retry decorator if available
        if TENACITY_AVAILABLE:
            self.generate_response = retry(
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=4, max=10)
            )(self.generate_response)
    
    def _initialize_providers(self):
        """Initialize available LLM providers"""
        # Try OpenAI
        try:
            if SECURE_CONFIG_AVAILABLE:
                api_key = config.openai_api_key
            else:
                api_key = os.getenv("OPENAI_API_KEY")
            if api_key and api_key.strip():
                try:
                    from openai import OpenAI
                    self.providers["openai"] = {
                        "client": OpenAI(api_key=api_key),
                        "model": "gpt-4-turbo-preview",
                        "name": "OpenAI GPT-4"
                    }
                    logger.info("OpenAI provider initialized")
                except ImportError:
                    logger.warning("OpenAI package not installed")
        except Exception as e:
            logger.error(f"Failed to initialize OpenAI: {e}")
            
        # Try Anthropic
        try:
            if SECURE_CONFIG_AVAILABLE:
                api_key = config.anthropic_api_key
            else:
                api_key = os.getenv("ANTHROPIC_API_KEY")
            if api_key and api_key.strip():
                try:
                    import anthropic
                    self.providers["anthropic"] = {
                        "client": anthropic.Anthropic(api_key=api_key),
                        "model": "claude-3-opus-20240229",
                        "name": "Anthropic Claude 3"
                    }
                    logger.info("Anthropic provider initialized")
                except ImportError:
                    logger.warning("Anthropic package not installed")
        except Exception as e:
            logger.error(f"Failed to initialize Anthropic: {e}")
            
        # Try Google Gemini
        try:
            if SECURE_CONFIG_AVAILABLE:
                api_key = config.google_api_key
            else:
                api_key = os.getenv("GOOGLE_API_KEY")
            if api_key and api_key.strip():
                try:
                    import google.generativeai as genai
                    genai.configure(api_key=api_key)
                    self.providers["gemini"] = {
                        "client": genai.GenerativeModel('gemini-pro'),
                        "model": "gemini-pro",
                        "name": "Google Gemini Pro"
                    }
                    logger.info("Gemini provider initialized")
                except ImportError:
                    logger.warning("Google Generative AI package not installed")
        except Exception as e:
            logger.error(f"Failed to initialize Gemini: {e}")

        # Set default provider if any providers are available
        if self.providers and not self.current_provider:
            # Prefer Anthropic (most reliable), then OpenAI, then Gemini
            preferred_order = ['anthropic', 'openai', 'gemini']
            for provider in preferred_order:
                if provider in self.providers:
                    self.current_provider = provider
                    logger.info(f"Set default provider to: {provider}")
                    break

        # Log initialization status
        logger.info(f"Initialized {len(self.providers)} LLM providers: {list(self.providers.keys())}")
        if self.current_provider:
            logger.info(f"Default provider set to: {self.current_provider}")
        else:
            logger.warning("No LLM providers available - will use fallback responses")

    def get_available_providers(self) -> Dict[str, bool]:
        """Get status of all providers"""
        return {name: True for name in self.providers.keys()}
    
    def get_current_provider(self) -> str:
        """Get name of current provider"""
        return self.current_provider or "None"
    
    def set_provider(self, provider_name: str) -> bool:
        """Set active provider"""
        if provider_name in self.providers:
            self.current_provider = provider_name
            return True
        return False
    
    def generate_response(self, prompt: str, provider: Optional[str] = None) -> str:
        """Generate response using specified or available provider with smart fallback"""
        if not self.providers:
            return self._fallback_response(prompt)
            
        provider_to_use = provider or self.current_provider
        
        # Try specified provider first
        if provider_to_use and provider_to_use in self.providers:
            try:
                response = self._call_provider(provider_to_use, prompt)
                if response and len(response.strip()) > 10:  # Valid response
                    return response
            except Exception as e:
                logger.warning(f"Provider {provider_to_use} failed: {e}")
        
        # Try all available providers
        for provider_name in self.providers.keys():
            if provider_name == provider_to_use:  # Skip already tried
                continue
            try:
                self.current_provider = provider_name
                response = self._call_provider(provider_name, prompt)
                if response and len(response.strip()) > 10:  # Valid response
                    return response
            except Exception as e:
                logger.warning(f"Provider {provider_name} failed: {e}")
                continue
        
        # All providers failed, use fallback
        return self._fallback_response(prompt)
    
    def _call_provider(self, provider: str, prompt: str) -> str:
        """Call specific provider with proper error handling"""
        if provider not in self.providers:
            return self._fallback_response(prompt)
            
        provider_info = self.providers[provider]
        
        try:
            if provider == "openai":
                response = provider_info["client"].chat.completions.create(
                    model=provider_info["model"],
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.7,
                    max_tokens=2000
                )
                return response.choices[0].message.content
                
            elif provider == "anthropic":
                response = provider_info["client"].messages.create(
                    model=provider_info["model"],
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=2000
                )
                return response.content[0].text
                
            elif provider == "gemini":
                response = provider_info["client"].generate_content(prompt)
                return response.text
                
        except Exception as e:
            logger.error(f"Provider {provider} error: {e}")
            
        return self._fallback_response(prompt)
    
    def _fallback_response(self, prompt: str) -> str:
        """Smart fallback responses when no AI available"""
        prompt_lower = prompt.lower()
        
        # Topic extraction fallback
        if "extract" in prompt_lower and "topic" in prompt_lower:
            return "I need help identifying the research topic. Please state it clearly, such as 'machine learning in healthcare' or 'climate change impacts'."
        
        # Gap analysis fallback
        if "gap" in prompt_lower and "analyze" in prompt_lower:
            return """Without AI assistance, I can identify:
• Temporal gaps (years with few publications)
• Venue concentration (limited publication venues)

For deeper analysis including methodological and theoretical gaps, an AI provider (OpenAI, Anthropic, or Google) API key is recommended."""
        
        # Query generation fallback
        if "query" in prompt_lower or "search" in prompt_lower:
            return "I'll search using basic query variations. For more sophisticated search strategies, consider enabling an AI provider."
        
        # General fallback
        return """I can help you search for papers and identify basic research gaps. 

What I can do without AI:
• Search academic databases (Semantic Scholar, Crossref)
• Find recent papers (last 5 years)
• Identify temporal research gaps
• Export your findings

For advanced features (smart topic extraction, detailed gap analysis), add an API key for OpenAI, Anthropic, or Google Gemini."""


class PaperValidator:
    """Validates papers to prevent fabrication"""
    
    @staticmethod
    def validate_paper(paper: Paper) -> bool:
        """Validate paper data for authenticity"""
        current_year = datetime.now().year
        
        # Check year range
        if not (1900 <= paper.year <= current_year):
            logger.warning(f"Invalid year {paper.year} for paper: {paper.title}")
            return False
        
        # Check title
        if len(paper.title) < 10 or len(paper.title) > 300:
            logger.warning(f"Suspicious title length: {paper.title}")
            return False
        
        # Check authors
        if not paper.authors or len(paper.authors) > 50:
            logger.warning(f"Suspicious author count: {paper.title}")
            return False
        
        # Check citations
        paper_age = current_year - paper.year
        if paper.citations > paper_age * 1000:
            logger.warning(f"Suspicious citation count: {paper.citations} for {paper.title}")
            return False
        
        return True


class SemanticScholarRateLimiter:
    """Rate limiter for Semantic Scholar API to ensure 1 request per second compliance"""

    def __init__(self, rate_limit_seconds: float = SEMANTIC_SCHOLAR_RATE_LIMIT):
        self.rate_limit = rate_limit_seconds
        self.last_request_time = 0
        self.lock = threading.Lock()

    def wait_if_needed(self, status_callback=None):
        """Wait if necessary to comply with rate limit"""
        with self.lock:
            current_time = time.time()
            time_since_last = current_time - self.last_request_time

            if time_since_last < self.rate_limit:
                wait_time = self.rate_limit - time_since_last
                if status_callback:
                    status_callback(f"⏳ Waiting for Semantic Scholar rate limit ({wait_time:.1f}s)...")
                logger.info(f"Rate limiting: waiting {wait_time:.1f} seconds for Semantic Scholar")
                time.sleep(wait_time)

            self.last_request_time = time.time()

    def exponential_backoff_wait(self, attempt: int, status_callback=None):
        """Exponential backoff for retries"""
        wait_time = RETRY_BACKOFF_BASE ** attempt
        if status_callback:
            status_callback(f"⏳ Rate limited - retrying in {wait_time}s (attempt {attempt + 1}/{MAX_RETRIES})...")
        logger.warning(f"Exponential backoff: waiting {wait_time} seconds (attempt {attempt + 1})")
        time.sleep(wait_time)


class AcademicSearchEngine:
    """Handles academic paper searches across multiple databases with proper rate limiting"""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'ResearchGapFinder/1.0 (Academic Research Tool)'
        })
        self.cache = {}
        self.semantic_scholar_limiter = SemanticScholarRateLimiter()
        self.status_callback = None  # Will be set by calling code for user feedback
        
        # Apply retry decorators if available
        if TENACITY_AVAILABLE:
            self.search_semantic_scholar = retry(
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=2, max=10)
            )(self.search_semantic_scholar)

            self.search_crossref = retry(
                stop=stop_after_attempt(3),
                wait=wait_exponential(multiplier=1, min=2, max=10)
            )(self.search_crossref)

    def search_semantic_scholar(self, query: str, limit: int = 20) -> List[Paper]:
        """Search Semantic Scholar API with proper rate limiting and retry logic"""
        cache_key = f"ss_{query}_{limit}"
        if cache_key in self.cache:
            cache_time, papers = self.cache[cache_key]
            if time.time() - cache_time < CACHE_DURATION:
                logger.info(f"Returning cached results for: {query}")
                return papers

        # Get API key from secure configuration
        if SECURE_CONFIG_AVAILABLE:
            api_key = config.semantic_scholar_api_key
        else:
            api_key = os.getenv('SEMANTIC_SCHOLAR_API_KEY')

        params = {
            "query": query,
            "limit": limit,
            "fields": "title,authors,year,venue,abstract,citationCount,externalIds,url",
            "offset": 0
        }

        headers = {
            'User-Agent': 'ResearchGapFinder/1.0 (Academic Research Tool)',
            'Accept': 'application/json'
        }

        # Add API key if available
        if api_key:
            headers['x-api-key'] = api_key

        # Implement retry logic with exponential backoff
        for attempt in range(MAX_RETRIES):
            try:
                # Apply rate limiting before each request
                self.semantic_scholar_limiter.wait_if_needed(self.status_callback)

                if self.status_callback:
                    self.status_callback(f"🔍 Searching Semantic Scholar... (attempt {attempt + 1})")

                response = self.session.get(
                    SEMANTIC_SCHOLAR_API,
                    params=params,
                    headers=headers,
                    timeout=15
                )

                if response.status_code == 429:  # Rate limit
                    logger.warning(f"Semantic Scholar rate limit hit on attempt {attempt + 1}")
                    if attempt < MAX_RETRIES - 1:
                        self.semantic_scholar_limiter.exponential_backoff_wait(attempt, self.status_callback)
                        continue
                    else:
                        logger.error("Max retries reached for Semantic Scholar rate limiting")
                        return []

                response.raise_for_status()
                data = response.json()
                break  # Success, exit retry loop

            except requests.exceptions.Timeout:
                logger.warning(f"Semantic Scholar timeout on attempt {attempt + 1}")
                if attempt < MAX_RETRIES - 1:
                    self.semantic_scholar_limiter.exponential_backoff_wait(attempt, self.status_callback)
                    continue
                else:
                    logger.error("Semantic Scholar request timed out after all retries")
                    return []
            except requests.exceptions.RequestException as e:
                logger.warning(f"Semantic Scholar request failed on attempt {attempt + 1}: {e}")
                if attempt < MAX_RETRIES - 1:
                    self.semantic_scholar_limiter.exponential_backoff_wait(attempt, self.status_callback)
                    continue
                else:
                    logger.error(f"Semantic Scholar request failed after all retries: {e}")
                    return []
            except Exception as e:
                logger.warning(f"Semantic Scholar error on attempt {attempt + 1}: {e}")
                if attempt < MAX_RETRIES - 1:
                    self.semantic_scholar_limiter.exponential_backoff_wait(attempt, self.status_callback)
                    continue
                else:
                    logger.error(f"Semantic Scholar search error after all retries: {e}")
                    return []

        # Process successful response
        papers = []
        for item in data.get("data", []):
            # Safely extract data
            try:
                paper = Paper(
                    title=item.get("title", "").strip(),
                    authors=[a.get("name", "Unknown") for a in item.get("authors", [])[:10]],
                    year=item.get("year") or 0,
                    venue=item.get("venue", "").strip() or "Preprint",
                    abstract=(item.get("abstract") or "")[:500],
                    citations=item.get("citationCount") or 0,
                    doi=item.get("externalIds", {}).get("DOI"),
                    url=item.get("url"),
                    source="Semantic Scholar"
                )

                if PaperValidator.validate_paper(paper):
                    papers.append(paper)
            except Exception as e:
                logger.warning(f"Failed to parse paper: {e}")
                continue

        self.cache[cache_key] = (time.time(), papers)
        logger.info(f"Found {len(papers)} valid papers from Semantic Scholar for: {query}")
        return papers
    
    def search_crossref(self, query: str, limit: int = 20) -> List[Paper]:
        """Search Crossref API with improved error handling"""
        cache_key = f"cr_{query}_{limit}"
        if cache_key in self.cache:
            cache_time, papers = self.cache[cache_key]
            if time.time() - cache_time < CACHE_DURATION:
                logger.info(f"Returning cached Crossref results for: {query}")
                return papers
        
        current_year = datetime.now().year
        params = {
            "query": query,
            "rows": limit,
            "select": "title,author,published-print,container-title,abstract,is-referenced-by-count,DOI,URL",
            "filter": f"from-pub-date:{current_year - 5}",  # Recent papers only
            "mailto": "research-gap-finder@example.com"  # Good practice for Crossref
        }
        
        headers = {
            'User-Agent': 'ResearchGapFinder/1.0 (Academic Research Tool; mailto:research@example.com)',
            'Accept': 'application/json'
        }
        
        try:
            response = self.session.get(
                CROSSREF_API,
                params=params,
                headers=headers,
                timeout=15
            )
            
            if response.status_code == 429:  # Rate limit
                logger.warning("Crossref rate limit hit, waiting...")
                time.sleep(2)
                raise Exception("Rate limit - will retry")
                
            response.raise_for_status()
            data = response.json()
            
            papers = []
            items = data.get("message", {}).get("items", [])
            
            for item in items:
                try:
                    # Safely extract year with multiple fallbacks
                    year = 0
                    
                    # Try published-print first
                    date_parts = item.get("published-print", {}).get("date-parts", [[]])
                    if date_parts and date_parts[0] and len(date_parts[0]) > 0:
                        year = date_parts[0][0]
                    
                    # Try published-online if no print date
                    if year == 0:
                        date_parts = item.get("published-online", {}).get("date-parts", [[]])
                        if date_parts and date_parts[0] and len(date_parts[0]) > 0:
                            year = date_parts[0][0]
                    
                    # Try created date as last resort
                    if year == 0:
                        date_parts = item.get("created", {}).get("date-parts", [[]])
                        if date_parts and date_parts[0] and len(date_parts[0]) > 0:
                            year = date_parts[0][0]
                    
                    # Extract title
                    title = ""
                    if item.get("title"):
                        title = item["title"][0] if isinstance(item["title"], list) else item["title"]
                    
                    # Extract venue
                    venue = ""
                    if item.get("container-title"):
                        venue = item["container-title"][0] if isinstance(item["container-title"], list) else item["container-title"]
                    
                    paper = Paper(
                        title=title.strip(),
                        authors=[f"{a.get('given', '')} {a.get('family', '')}".strip() 
                                for a in item.get("author", [])[:10]],
                        year=year,
                        venue=venue.strip() or "Preprint",
                        abstract=(item.get("abstract", ""))[:500],
                        citations=item.get("is-referenced-by-count", 0),
                        doi=item.get("DOI"),
                        url=item.get("URL"),
                        source="Crossref"
                    )
                    
                    if PaperValidator.validate_paper(paper):
                        papers.append(paper)
                        
                except Exception as e:
                    logger.warning(f"Failed to parse Crossref paper: {e}")
                    continue
            
            self.cache[cache_key] = (time.time(), papers)
            logger.info(f"Found {len(papers)} valid papers from Crossref for: {query}")
            return papers
            
        except requests.exceptions.Timeout:
            logger.error("Crossref request timed out")
            return []
        except requests.exceptions.RequestException as e:
            logger.error(f"Crossref request failed: {e}")
            return []
        except Exception as e:
            logger.error(f"Crossref search error: {e}")
            return []
    
    def search_all_sources(self, query: str, status_callback=None) -> List[Paper]:
        """Search all available sources with proper rate limiting for Semantic Scholar"""
        all_papers = []
        search_errors = []

        # Set status callback for user feedback
        self.status_callback = status_callback

        # Show progress in logs
        logger.info(f"Starting search for: {query}")

        # Search Semantic Scholar first (sequential due to rate limiting)
        if status_callback:
            status_callback("🔍 Searching Semantic Scholar (rate limited)...")

        try:
            semantic_papers = self.search_semantic_scholar(query)
            all_papers.extend(semantic_papers)
            logger.info(f"✓ Semantic Scholar returned {len(semantic_papers)} papers")
            if status_callback:
                status_callback(f"✅ Semantic Scholar: {len(semantic_papers)} papers found")
        except Exception as e:
            error_msg = f"✗ Semantic Scholar failed: {str(e)}"
            logger.error(error_msg)
            search_errors.append(error_msg)
            if status_callback:
                status_callback(f"⚠️ Semantic Scholar: {str(e)}")

        # Search Crossref concurrently (no rate limiting needed)
        if status_callback:
            status_callback("🔍 Searching Crossref...")

        try:
            crossref_papers = self.search_crossref(query)
            all_papers.extend(crossref_papers)
            logger.info(f"✓ Crossref returned {len(crossref_papers)} papers")
            if status_callback:
                status_callback(f"✅ Crossref: {len(crossref_papers)} papers found")
        except Exception as e:
            error_msg = f"✗ Crossref failed: {str(e)}"
            logger.error(error_msg)
            search_errors.append(error_msg)
            if status_callback:
                status_callback(f"⚠️ Crossref: {str(e)}")
        
        # Log search summary
        if search_errors:
            logger.warning(f"Search completed with errors: {'; '.join(search_errors)}")
        else:
            logger.info(f"Search completed successfully, total papers: {len(all_papers)}")
        
        # Deduplicate papers based on title similarity
        seen_titles = set()
        unique_papers = []
        
        for paper in all_papers:
            # Create normalized title for comparison
            normalized_title = re.sub(r'[^\w\s]', '', paper.title.lower()).strip()
            
            # Check for similar titles (to catch slight variations)
            is_duplicate = False
            for seen_title in seen_titles:
                # Calculate simple similarity
                if normalized_title == seen_title:
                    is_duplicate = True
                    break
                # Check if one title contains the other (common with subtitles)
                if len(normalized_title) > 20 and len(seen_title) > 20:
                    if normalized_title in seen_title or seen_title in normalized_title:
                        is_duplicate = True
                        break
            
            if not is_duplicate:
                seen_titles.add(normalized_title)
                unique_papers.append(paper)
        
        # Calculate relevance scores
        current_year = datetime.now().year
        for paper in unique_papers:
            # Recency score (0-1, where 1 is current year)
            recency_score = max(0, 1 - (current_year - paper.year) / 10)
            
            # Citation score (normalized log scale)
            citation_score = min(1, (paper.citations + 1) ** 0.25 / 10)
            
            # Combined relevance score
            paper.relevance_score = (recency_score * 0.4) + (citation_score * 0.6)
        
        # Sort by relevance
        unique_papers.sort(key=lambda p: p.relevance_score, reverse=True)
        
        # Return top papers
        result = unique_papers[:MAX_PAPERS_PER_SOURCE]
        logger.info(f"Returning {len(result)} unique papers after deduplication")
        return result


class ResearchGapAnalyzer:
    """Main research assistant engine for gap analysis with enhanced user guidance"""

    def __init__(self):
        self.llm = LLMProviderManager()
        self.searcher = AcademicSearchEngine()
        self.conversation_phase = ConversationPhase.INITIALIZATION
        self.current_topic = None
        self.user_expertise = None
        self.user_interests = []
        self.recent_papers = []
        self.paper_scores = {}  # paper_id -> PaperQualityScore
        self.identified_gaps = []
        self.conversation_history = []
        self.progress_status = ""
        self.next_steps = []
        
    def extract_research_topic(self, user_input: str) -> Optional[str]:
        """Extract research topic from user input with smart detection"""
        # Enhanced pattern matching
        patterns = [
            r"(?:research|study|investigate|explore|analyze|examine) (?:on |about |in |into )?(.+?)(?:\?|\.|\!|$)",
            r"(?:papers|publications|articles|literature) (?:on |about |in |regarding )?(.+?)(?:\?|\.|\!|$)",
            r"(?:interested in|working on|focusing on|looking at) (.+?)(?:\?|\.|\!|$)",
            r"(?:topic|area|field|domain) (?:is |of )(.+?)(?:\?|\.|\!|$)",
            r"(.+?) (?:research|studies|papers|literature)",
            r"(?:find|search|look for) (?:papers on |about |regarding )?(.+?)(?:\?|\.|\!|$)",
            r"(.+?) (?:in|for|and) (?:healthcare|education|industry|science|technology)",
            r"(?:I want to|I'm|I am) (?:research|study|explore) (.+?)(?:\?|\.|\!|$)"
        ]
        
        lower_input = user_input.lower().strip()
        
        # Direct topic in quotes
        quote_match = re.search(r'"([^"]+)"', lower_input)
        if quote_match:
            return quote_match.group(1).strip()
        
        # Try patterns
        for pattern in patterns:
            match = re.search(pattern, lower_input)
            if match:
                topic = match.group(1).strip()
                # Clean up common endings
                topic = re.sub(r'\b(?:papers?|research|studies|literature|please|thanks)\b', '', topic).strip()
                topic = re.sub(r'[^\w\s-]', '', topic).strip()
                if len(topic) > 3 and len(topic.split()) <= 8:  # Reasonable topic length
                    return topic
        
        # Try LLM extraction if available
        if self.llm.providers:
            prompt = f"""You are a research topic extraction expert. Extract the main research topic from this user input.

Rules:
1. Return ONLY the core research topic (2-6 words)
2. Remove filler words like "papers on", "research about", etc.
3. If no clear research topic exists, return "None"
4. Focus on academic/scientific topics only
5. Keep it concise and specific

Examples:
- "I want to research machine learning in healthcare" → "machine learning healthcare"
- "Find papers about climate change impacts" → "climate change impacts"
- "Hello how are you" → "None"

User input: "{user_input}"

Topic:"""
            
            try:
                response = self.llm.generate_response(prompt)
                topic = response.strip().strip('"').strip("'")
                
                if topic.lower() != "none" and 3 < len(topic) < 100:
                    return topic
            except Exception as e:
                logger.warning(f"LLM topic extraction failed: {e}")
        
        return None
    
    def find_recent_papers(self, topic: str, status_callback=None) -> List[Paper]:
        """Find recent papers with intelligent query generation and rate limiting"""
        # Generate smart search queries
        search_queries = self._generate_search_queries(topic)

        if status_callback:
            status_callback(f"🔍 Generated {len(search_queries)} search queries for '{topic}'")

        all_papers = []
        seen_titles = set()

        for i, query in enumerate(search_queries[:3]):  # Limit to 3 queries
            try:
                if status_callback:
                    status_callback(f"📚 Searching with query {i+1}/{min(len(search_queries), 3)}: '{query[:50]}...'")

                # Pass status callback to search engine
                papers = self.searcher.search_all_sources(query, status_callback)

                # Deduplicate as we go
                new_papers = 0
                for paper in papers:
                    title_key = paper.title.lower().strip()
                    if title_key not in seen_titles:
                        seen_titles.add(title_key)
                        all_papers.append(paper)
                        new_papers += 1

                if status_callback:
                    status_callback(f"📄 Query {i+1} added {new_papers} new papers (total: {len(all_papers)})")

                # Add delay between queries for additional rate limiting safety
                if i < len(search_queries) - 1:
                    if status_callback:
                        status_callback("⏳ Brief pause between queries...")
                    time.sleep(1.5)  # Increased delay for safety

            except Exception as e:
                logger.error(f"Search failed for query '{query}': {e}")
                if status_callback:
                    status_callback(f"⚠️ Query {i+1} failed: {str(e)}")
        
        # Filter for recency (last 5 years)
        current_year = datetime.now().year
        recent_papers = [p for p in all_papers if p.year >= current_year - 5]
        
        # Sort by relevance
        recent_papers.sort(key=lambda p: (p.year, p.citations), reverse=True)
        
        # Log search results
        logger.info(f"Found {len(recent_papers)} recent papers for topic: {topic}")
        
        return recent_papers[:20]
    
    def _generate_search_queries(self, topic: str) -> List[str]:
        """Generate intelligent search queries for comprehensive coverage"""
        queries = [topic]  # Always include the base topic
        
        # Add variations based on common patterns
        if len(topic.split()) <= 3:  # Short topic
            queries.extend([
                f"{topic} recent advances",
                f"{topic} systematic review",
                f"{topic} survey"
            ])
        
        # Try LLM query generation if available
        if self.llm.providers:
            prompt = f"""Generate 3 different search queries to find academic papers on: "{topic}"

Requirements:
1. Queries should be 2-6 words each
2. Cover different aspects of the topic
3. Use academic terminology
4. Avoid overly broad or narrow queries
5. Include one query for recent reviews/surveys

Topic: {topic}

Return ONLY a JSON array of 3 queries, like: ["query 1", "query 2", "query 3"]"""
            
            try:
                response = self.llm.generate_response(prompt)
                # Extract JSON array
                json_match = re.search(r'\[.*?\]', response)
                if json_match:
                    llm_queries = json.loads(json_match.group())
                    # Validate and add queries
                    for q in llm_queries[:3]:
                        if isinstance(q, str) and 2 <= len(q.split()) <= 8:
                            queries.append(q)
            except Exception as e:
                logger.warning(f"LLM query generation failed: {e}")
        
        # Remove duplicates while preserving order
        seen = set()
        unique_queries = []
        for q in queries:
            q_lower = q.lower().strip()
            if q_lower not in seen:
                seen.add(q_lower)
                unique_queries.append(q)
        
        return unique_queries[:4]  # Return max 4 queries

    def grade_paper_quality(self, paper: Paper) -> PaperQualityScore:
        """Comprehensive paper quality assessment with multiple criteria"""
        try:
            # Calculate individual scores
            methodology_score = self._assess_methodology(paper)
            citation_impact_score = self._assess_citation_impact(paper)
            novelty_score = self._assess_novelty(paper)
            clarity_score = self._assess_clarity(paper)
            applicability_score = self._assess_applicability(paper)

            # Calculate overall score (weighted average)
            weights = {
                'methodology': 0.25,
                'citation_impact': 0.20,
                'novelty': 0.25,
                'clarity': 0.15,
                'applicability': 0.15
            }

            overall_score = (
                methodology_score * weights['methodology'] +
                citation_impact_score * weights['citation_impact'] +
                novelty_score * weights['novelty'] +
                clarity_score * weights['clarity'] +
                applicability_score * weights['applicability']
            )

            # Convert to letter grade
            overall_grade = self._score_to_grade(overall_score)

            # Generate explanation and strengths/weaknesses
            explanation, strengths, weaknesses = self._generate_paper_assessment(
                paper, methodology_score, citation_impact_score, novelty_score,
                clarity_score, applicability_score
            )

            return PaperQualityScore(
                overall_grade=overall_grade,
                overall_score=round(overall_score, 1),
                methodology_score=round(methodology_score, 1),
                citation_impact_score=round(citation_impact_score, 1),
                novelty_score=round(novelty_score, 1),
                clarity_score=round(clarity_score, 1),
                applicability_score=round(applicability_score, 1),
                explanation=explanation,
                strengths=strengths,
                weaknesses=weaknesses
            )

        except Exception as e:
            logger.error(f"Error grading paper {paper.title}: {e}")
            # Return default grade
            return PaperQualityScore(
                overall_grade=PaperGrade.C,
                overall_score=5.0,
                methodology_score=5.0,
                citation_impact_score=5.0,
                novelty_score=5.0,
                clarity_score=5.0,
                applicability_score=5.0,
                explanation="Unable to assess paper quality due to processing error.",
                strengths=["Paper included in academic database"],
                weaknesses=["Assessment could not be completed"]
            )

    def _assess_methodology(self, paper: Paper) -> float:
        """Assess methodology rigor (0-10 scale)"""
        score = 5.0  # Base score

        # Check for methodology keywords in abstract
        methodology_keywords = [
            'experiment', 'empirical', 'statistical', 'quantitative', 'qualitative',
            'survey', 'case study', 'longitudinal', 'cross-sectional', 'randomized',
            'controlled', 'systematic', 'meta-analysis', 'validation', 'evaluation'
        ]

        abstract_lower = paper.abstract.lower()
        keyword_count = sum(1 for keyword in methodology_keywords if keyword in abstract_lower)
        score += min(keyword_count * 0.5, 3.0)  # Up to 3 points for methodology keywords

        # Venue quality (if available)
        if paper.venue and len(paper.venue) > 3:
            if any(term in paper.venue.lower() for term in ['ieee', 'acm', 'nature', 'science']):
                score += 1.5
            elif any(term in paper.venue.lower() for term in ['conference', 'journal']):
                score += 0.5

        return min(score, 10.0)

    def _assess_citation_impact(self, paper: Paper) -> float:
        """Assess citation impact relative to paper age (0-10 scale)"""
        current_year = datetime.now().year
        paper_age = max(current_year - paper.year, 1)

        # Calculate citations per year
        citations_per_year = paper.citations / paper_age

        # Score based on citations per year
        if citations_per_year >= 50:
            return 10.0
        elif citations_per_year >= 20:
            return 8.5
        elif citations_per_year >= 10:
            return 7.0
        elif citations_per_year >= 5:
            return 6.0
        elif citations_per_year >= 2:
            return 5.0
        elif citations_per_year >= 1:
            return 4.0
        else:
            return max(2.0, 3.0 - paper_age * 0.2)  # Newer papers get benefit of doubt

    def _assess_novelty(self, paper: Paper) -> float:
        """Assess novelty and innovation (0-10 scale)"""
        score = 5.0  # Base score

        novelty_keywords = [
            'novel', 'new', 'innovative', 'first', 'breakthrough', 'pioneering',
            'unprecedented', 'original', 'unique', 'cutting-edge', 'state-of-the-art'
        ]

        title_lower = paper.title.lower()
        abstract_lower = paper.abstract.lower()

        # Check title for novelty indicators
        title_novelty = sum(1 for keyword in novelty_keywords if keyword in title_lower)
        score += min(title_novelty * 1.0, 2.0)

        # Check abstract for novelty indicators
        abstract_novelty = sum(1 for keyword in novelty_keywords if keyword in abstract_lower)
        score += min(abstract_novelty * 0.3, 2.0)

        # Recent papers get slight novelty boost
        current_year = datetime.now().year
        if paper.year >= current_year - 1:
            score += 1.0
        elif paper.year >= current_year - 3:
            score += 0.5

        return min(score, 10.0)

    def _assess_clarity(self, paper: Paper) -> float:
        """Assess clarity of presentation (0-10 scale)"""
        score = 5.0  # Base score

        # Check title clarity (not too long, not too short)
        title_length = len(paper.title)
        if 30 <= title_length <= 120:
            score += 1.0
        elif title_length < 20 or title_length > 150:
            score -= 1.0

        # Check abstract quality
        if paper.abstract:
            abstract_length = len(paper.abstract)
            if 100 <= abstract_length <= 500:
                score += 1.0

            # Check for clear structure indicators
            structure_keywords = [
                'background', 'objective', 'method', 'result', 'conclusion',
                'purpose', 'approach', 'finding', 'implication'
            ]
            abstract_lower = paper.abstract.lower()
            structure_count = sum(1 for keyword in structure_keywords if keyword in abstract_lower)
            score += min(structure_count * 0.3, 2.0)

        # Author count (too many authors might indicate less clarity)
        if len(paper.authors) <= 6:
            score += 0.5
        elif len(paper.authors) > 15:
            score -= 0.5

        return min(score, 10.0)

    def _assess_applicability(self, paper: Paper) -> float:
        """Assess practical applicability (0-10 scale)"""
        score = 5.0  # Base score

        application_keywords = [
            'application', 'practical', 'implementation', 'real-world', 'industry',
            'deployment', 'system', 'tool', 'framework', 'platform', 'solution',
            'case study', 'evaluation', 'performance', 'efficiency'
        ]

        title_lower = paper.title.lower()
        abstract_lower = paper.abstract.lower()

        # Check for application indicators
        title_apps = sum(1 for keyword in application_keywords if keyword in title_lower)
        abstract_apps = sum(1 for keyword in application_keywords if keyword in abstract_lower)

        score += min(title_apps * 1.0, 2.0)
        score += min(abstract_apps * 0.2, 2.0)

        # Check for specific domains that indicate applicability
        domain_keywords = [
            'healthcare', 'medical', 'finance', 'education', 'transportation',
            'energy', 'environment', 'security', 'manufacturing', 'agriculture'
        ]

        domain_mentions = sum(1 for keyword in domain_keywords
                            if keyword in title_lower or keyword in abstract_lower)
        score += min(domain_mentions * 0.5, 1.0)

        return min(score, 10.0)

    def _score_to_grade(self, score: float) -> PaperGrade:
        """Convert numerical score to letter grade"""
        if score >= 9.5:
            return PaperGrade.A_PLUS
        elif score >= 9.0:
            return PaperGrade.A
        elif score >= 8.5:
            return PaperGrade.A_MINUS
        elif score >= 8.0:
            return PaperGrade.B_PLUS
        elif score >= 7.0:
            return PaperGrade.B
        elif score >= 6.5:
            return PaperGrade.B_MINUS
        elif score >= 6.0:
            return PaperGrade.C_PLUS
        elif score >= 5.0:
            return PaperGrade.C
        elif score >= 4.0:
            return PaperGrade.C_MINUS
        elif score >= 3.0:
            return PaperGrade.D
        else:
            return PaperGrade.F

    def _generate_paper_assessment(self, paper: Paper, methodology: float, citation: float,
                                 novelty: float, clarity: float, applicability: float) -> Tuple[str, List[str], List[str]]:
        """Generate detailed assessment explanation with strengths and weaknesses"""
        strengths = []
        weaknesses = []

        # Analyze strengths
        if methodology >= 8.0:
            strengths.append("Strong methodological approach")
        if citation >= 7.0:
            strengths.append("High citation impact for its age")
        if novelty >= 8.0:
            strengths.append("Highly novel and innovative")
        if clarity >= 8.0:
            strengths.append("Clear and well-structured presentation")
        if applicability >= 8.0:
            strengths.append("Strong practical applicability")

        # Analyze weaknesses
        if methodology < 5.0:
            weaknesses.append("Limited methodological rigor")
        if citation < 4.0:
            weaknesses.append("Low citation impact")
        if novelty < 5.0:
            weaknesses.append("Limited novelty or innovation")
        if clarity < 5.0:
            weaknesses.append("Unclear presentation or structure")
        if applicability < 5.0:
            weaknesses.append("Limited practical applicability")

        # Generate explanation
        overall_score = (methodology + citation + novelty + clarity + applicability) / 5

        if overall_score >= 8.5:
            explanation = "Excellent paper with strong contributions across multiple dimensions."
        elif overall_score >= 7.5:
            explanation = "High-quality paper with solid methodology and significant impact."
        elif overall_score >= 6.5:
            explanation = "Good paper with notable contributions and clear presentation."
        elif overall_score >= 5.5:
            explanation = "Adequate paper with some valuable insights but room for improvement."
        else:
            explanation = "Paper has limitations that affect its overall quality and impact."

        return explanation, strengths, weaknesses
    
    def analyze_research_gaps(self, papers: List[Paper]) -> List[ResearchGap]:
        """Analyze papers to identify research gaps with enhanced intelligence"""
        if not papers:
            return []
        
        gaps = []
        current_year = datetime.now().year
        
        # Basic temporal analysis (always available)
        year_counts = {}
        venue_counts = {}
        
        for paper in papers:
            year_counts[paper.year] = year_counts.get(paper.year, 0) + 1
            if paper.venue:
                venue_counts[paper.venue] = venue_counts.get(paper.venue, 0) + 1
        
        # Temporal gaps
        for year in range(current_year - 5, current_year + 1):
            if year_counts.get(year, 0) < 2:
                gaps.append(ResearchGap(
                    gap_type="temporal",
                    description=f"Limited research activity in {year} with only {year_counts.get(year, 0)} publications",
                    supporting_evidence=[
                        f"Only {year_counts.get(year, 0)} papers found from {year}",
                        f"Average publications per year: {len(papers) / 5:.1f}"
                    ],
                    potential_impact="Recent developments and current trends may be understudied",
                    suggested_approach="Conduct updated research incorporating latest developments and technologies",
                    confidence_score=0.7
                ))
        
        # Venue diversity gap
        if len(venue_counts) < 3:
            gaps.append(ResearchGap(
                gap_type="venue_diversity",
                description="Limited venue diversity suggests narrow research community engagement",
                supporting_evidence=[
                    f"Papers concentrated in only {len(venue_counts)} venues",
                    f"Top venue: {max(venue_counts, key=venue_counts.get) if venue_counts else 'Unknown'}"
                ],
                potential_impact="Potential echo chamber effect limiting diverse perspectives",
                suggested_approach="Target interdisciplinary venues and conferences for broader impact",
                confidence_score=0.6
            ))
        
        # Enhanced LLM analysis if available
        if self.llm.providers and papers:
            # Prepare comprehensive paper analysis
            paper_data = []
            for p in papers[:15]:  # Limit for context
                abstract_text = p.abstract[:200] if p.abstract else "No abstract available"
                paper_info = {
                    "title": p.title,
                    "year": p.year,
                    "venue": p.venue or "Unknown venue",
                    "citations": p.citations,
                    "abstract_snippet": abstract_text
                }
                paper_data.append(paper_info)
            
            papers_json = json.dumps(paper_data, indent=2)
            
            prompt = f"""You are an expert research gap analyst. Analyze these {len(papers)} papers on "{self.current_topic}" to identify significant research gaps.

Papers Data:
{papers_json}

Identify 3-5 SPECIFIC and ACTIONABLE research gaps. For each gap:

1. Classify the gap type:
   - methodological: Missing research methods or approaches
   - theoretical: Lacking theoretical frameworks or models
   - application: Unexplored practical applications
   - interdisciplinary: Missing connections to other fields
   - technological: Not utilizing recent tech advances
   - geographical: Limited geographical coverage
   - demographic: Understudied populations

2. Provide concrete evidence from the papers
3. Explain the significance and potential impact
4. Suggest specific, actionable research approaches

Format your response as a JSON array with objects containing:
{{
  "gap_type": "type",
  "description": "Clear, specific description",
  "supporting_evidence": ["Evidence 1", "Evidence 2"],
  "potential_impact": "Why this matters",
  "suggested_approach": "Specific research approach",
  "key_questions": ["Question 1", "Question 2"]
}}

Be specific and avoid generic gaps. Focus on what's truly missing based on these papers."""
            
            try:
                response = self.llm.generate_response(prompt)
                
                # Extract JSON with better error handling
                json_match = re.search(r'\[\s*\{.*?\}\s*\]', response, re.DOTALL)
                if json_match:
                    gaps_data = json.loads(json_match.group())
                    
                    for gap_info in gaps_data[:5]:  # Max 5 gaps
                        # Validate gap data
                        if all(key in gap_info for key in ["gap_type", "description", "supporting_evidence"]):
                            # Enhanced gap creation with new features
                            gap = ResearchGap(
                                gap_type=gap_info.get("gap_type", "unknown"),
                                description=gap_info.get("description", ""),
                                supporting_evidence=gap_info.get("supporting_evidence", [])[:3],
                                potential_impact=gap_info.get("potential_impact", "Significant research opportunity"),
                                suggested_approach=gap_info.get("suggested_approach", "Further investigation needed"),
                                confidence_score=0.85,
                                novelty_grade=self._assess_gap_novelty(gap_info),
                                impact_grade=self._assess_gap_impact(gap_info),
                                research_suggestions=self._generate_research_suggestions(gap_info),
                                methodology_suggestions=self._generate_methodology_suggestions(gap_info),
                                required_expertise=self._determine_required_expertise(gap_info),
                                estimated_timeline=self._estimate_timeline(gap_info),
                                required_resources=self._identify_required_resources(gap_info)
                            )

                            # Add key questions to description if available
                            if "key_questions" in gap_info:
                                gap.description += "\n\nKey questions: " + "; ".join(gap_info["key_questions"][:2])

                            gaps.append(gap)
                            
            except Exception as e:
                logger.error(f"LLM gap analysis failed: {e}")
                # Add fallback gap
                gaps.append(ResearchGap(
                    gap_type="analysis_limitation",
                    description="Advanced gap analysis unavailable - consider manual literature review",
                    supporting_evidence=["AI analysis failed", "Manual review recommended"],
                    potential_impact="May miss nuanced research opportunities",
                    suggested_approach="Conduct systematic literature review with domain experts",
                    confidence_score=0.5
                ))
        
        # Sort by confidence and limit
        gaps.sort(key=lambda g: g.confidence_score, reverse=True)
        return gaps[:5]

    def _assess_gap_novelty(self, gap_info: dict) -> str:
        """Assess novelty level of research gap"""
        description = gap_info.get("description", "").lower()
        gap_type = gap_info.get("gap_type", "").lower()

        # High novelty indicators
        high_novelty_keywords = ["unprecedented", "first", "novel", "breakthrough", "cutting-edge", "unexplored"]
        if any(keyword in description for keyword in high_novelty_keywords):
            return "High"

        # Interdisciplinary and technological gaps tend to be high novelty
        if gap_type in ["interdisciplinary", "technological"]:
            return "High"

        # Medium novelty indicators
        medium_novelty_keywords = ["limited", "understudied", "emerging", "recent"]
        if any(keyword in description for keyword in medium_novelty_keywords):
            return "Medium"

        # Methodological gaps tend to be medium novelty
        if gap_type in ["methodological", "application"]:
            return "Medium"

        return "Low"

    def _assess_gap_impact(self, gap_info: dict) -> str:
        """Assess potential impact of research gap"""
        impact_text = gap_info.get("potential_impact", "").lower()
        description = gap_info.get("description", "").lower()

        # High impact indicators
        high_impact_keywords = ["significant", "major", "critical", "transformative", "breakthrough", "paradigm"]
        if any(keyword in impact_text or keyword in description for keyword in high_impact_keywords):
            return "High"

        # Application and interdisciplinary gaps tend to have high impact
        gap_type = gap_info.get("gap_type", "").lower()
        if gap_type in ["application", "interdisciplinary", "technological"]:
            return "High"

        # Medium impact indicators
        medium_impact_keywords = ["important", "valuable", "useful", "beneficial", "improve"]
        if any(keyword in impact_text or keyword in description for keyword in medium_impact_keywords):
            return "Medium"

        return "Medium"  # Default to medium

    def _generate_research_suggestions(self, gap_info: dict) -> List[str]:
        """Generate specific research suggestions"""
        gap_type = gap_info.get("gap_type", "").lower()
        description = gap_info.get("description", "")

        suggestions = []

        if gap_type == "methodological":
            suggestions.extend([
                "Develop and validate new research methodologies",
                "Conduct comparative studies of existing approaches",
                "Create standardized evaluation frameworks"
            ])
        elif gap_type == "theoretical":
            suggestions.extend([
                "Develop comprehensive theoretical frameworks",
                "Conduct systematic literature reviews",
                "Propose new conceptual models"
            ])
        elif gap_type == "application":
            suggestions.extend([
                "Design and implement pilot studies",
                "Develop proof-of-concept prototypes",
                "Conduct real-world case studies"
            ])
        elif gap_type == "interdisciplinary":
            suggestions.extend([
                "Establish cross-disciplinary research collaborations",
                "Organize interdisciplinary workshops and conferences",
                "Develop integrated research frameworks"
            ])
        else:
            suggestions.extend([
                "Conduct comprehensive literature review",
                "Design empirical studies to address the gap",
                "Develop new research methodologies"
            ])

        return suggestions[:3]  # Return top 3 suggestions

    def _generate_methodology_suggestions(self, gap_info: dict) -> List[str]:
        """Generate methodology suggestions for research gap"""
        gap_type = gap_info.get("gap_type", "").lower()

        methodologies = []

        if gap_type == "methodological":
            methodologies.extend([
                "Mixed-methods research design",
                "Systematic review and meta-analysis",
                "Experimental validation studies"
            ])
        elif gap_type == "theoretical":
            methodologies.extend([
                "Grounded theory approach",
                "Conceptual framework development",
                "Delphi study with experts"
            ])
        elif gap_type == "application":
            methodologies.extend([
                "Action research methodology",
                "Case study research",
                "Design science research"
            ])
        elif gap_type == "interdisciplinary":
            methodologies.extend([
                "Transdisciplinary research approach",
                "Collaborative participatory research",
                "Systems thinking methodology"
            ])
        else:
            methodologies.extend([
                "Exploratory research design",
                "Survey and interview methods",
                "Longitudinal study approach"
            ])

        return methodologies[:3]

    def _determine_required_expertise(self, gap_info: dict) -> UserExpertiseLevel:
        """Determine required expertise level for research gap"""
        gap_type = gap_info.get("gap_type", "").lower()
        description = gap_info.get("description", "").lower()

        # Expert level indicators
        expert_keywords = ["complex", "advanced", "sophisticated", "cutting-edge", "breakthrough"]
        if any(keyword in description for keyword in expert_keywords):
            return UserExpertiseLevel.EXPERT

        # Interdisciplinary and theoretical gaps often require advanced expertise
        if gap_type in ["interdisciplinary", "theoretical", "technological"]:
            return UserExpertiseLevel.ADVANCED

        # Application gaps can often be tackled by intermediate researchers
        if gap_type in ["application", "methodological"]:
            return UserExpertiseLevel.INTERMEDIATE

        return UserExpertiseLevel.INTERMEDIATE  # Default

    def _estimate_timeline(self, gap_info: dict) -> str:
        """Estimate timeline for addressing research gap"""
        gap_type = gap_info.get("gap_type", "").lower()
        description = gap_info.get("description", "").lower()

        # Long-term projects
        if gap_type in ["theoretical", "interdisciplinary"] or "comprehensive" in description:
            return "12-24 months"

        # Medium-term projects
        if gap_type in ["methodological", "technological"]:
            return "6-12 months"

        # Shorter-term projects
        if gap_type in ["application", "temporal"]:
            return "3-6 months"

        return "6-12 months"  # Default

    def _identify_required_resources(self, gap_info: dict) -> List[str]:
        """Identify required resources for research gap"""
        gap_type = gap_info.get("gap_type", "").lower()

        resources = []

        if gap_type == "methodological":
            resources.extend(["Statistical software", "Research participants", "Data collection tools"])
        elif gap_type == "theoretical":
            resources.extend(["Literature databases", "Expert consultations", "Analysis software"])
        elif gap_type == "application":
            resources.extend(["Development tools", "Testing environments", "User groups"])
        elif gap_type == "interdisciplinary":
            resources.extend(["Cross-domain experts", "Collaboration platforms", "Integration tools"])
        elif gap_type == "technological":
            resources.extend(["Computing resources", "Software licenses", "Technical expertise"])
        else:
            resources.extend(["Research databases", "Analysis tools", "Expert guidance"])

        return resources[:4]  # Return top 4 resources
    
    def process_message(self, user_input: str) -> str:
        """Enhanced process user input with step-by-step guidance"""
        self.conversation_history.append({"role": "user", "content": user_input})

        lower_input = user_input.lower()

        # Handle expertise level setting
        if any(phrase in lower_input for phrase in ["beginner", "intermediate", "advanced", "expert"]):
            return self._handle_expertise_setting(user_input)

        # Handle show more requests
        if any(phrase in lower_input for phrase in ["show more papers", "show all papers", "more papers"]):
            return self._handle_show_more_papers()

        if any(phrase in lower_input for phrase in ["show more gaps", "show all gaps", "more gaps"]):
            return self._handle_show_more_gaps()

        if any(phrase in lower_input for phrase in ["show more graded papers", "show all graded papers", "more graded papers"]):
            return self._handle_show_more_graded_papers()

        # Handle paper grading requests
        if any(phrase in lower_input for phrase in ["grade papers", "evaluate papers", "assess papers", "paper quality"]):
            return self._handle_paper_grading()

        # Check for paper search intent
        if any(phrase in lower_input for phrase in ["find papers", "search papers", "recent papers", "show papers"]):
            return self._handle_paper_search(user_input)

        # Check for gap analysis intent
        elif any(phrase in lower_input for phrase in ["analyze gaps", "research gaps", "find gaps", "what gaps"]):
            return self._handle_gap_analysis()

        # Handle domain refinement
        elif self.conversation_phase == ConversationPhase.DOMAIN_REFINEMENT:
            return self._handle_domain_refinement(user_input)

        # Try to extract topic
        else:
            topic = self.extract_research_topic(user_input)
            if topic:
                self.current_topic = topic
                self.conversation_phase = ConversationPhase.TOPIC_EXPLORATION
                return self._handle_topic_introduction(topic)
            else:
                return self._handle_general_query(user_input)

    def _handle_expertise_setting(self, user_input: str) -> str:
        """Handle user expertise level setting"""
        lower_input = user_input.lower()

        if "beginner" in lower_input:
            self.user_expertise = UserExpertiseLevel.BEGINNER
            expertise_msg = "beginner"
        elif "intermediate" in lower_input:
            self.user_expertise = UserExpertiseLevel.INTERMEDIATE
            expertise_msg = "intermediate"
        elif "advanced" in lower_input:
            self.user_expertise = UserExpertiseLevel.ADVANCED
            expertise_msg = "advanced"
        elif "expert" in lower_input:
            self.user_expertise = UserExpertiseLevel.EXPERT
            expertise_msg = "expert"
        else:
            return "I didn't catch your expertise level. Please specify: beginner, intermediate, advanced, or expert."

        self.conversation_phase = ConversationPhase.TOPIC_EXPLORATION

        return f"""### ✅ Expertise Level Set: {expertise_msg.title()}

Perfect! I've noted that you're at the **{expertise_msg}** level. This will help me:

• **Tailor research recommendations** to your experience level
• **Suggest appropriate methodologies** for your skill set
• **Identify gaps** that match your capabilities
• **Recommend resources** suitable for your expertise

**Next Step:** Tell me about your research area of interest. For example:
• "I'm interested in machine learning applications in healthcare"
• "I want to explore sustainable energy technologies"
• "Show me recent work in natural language processing"

What research topic would you like to explore?"""

    def _handle_paper_grading(self) -> str:
        """Handle paper quality assessment request"""
        if not self.recent_papers:
            return """### 📊 Paper Quality Assessment

I'd love to grade papers for you! However, I need to search for papers first.

**To get started:**
1. Tell me your research topic (e.g., "machine learning in healthcare")
2. I'll search for recent papers
3. Then I'll provide comprehensive quality grades for each paper

What research area should I search for papers to grade?"""

        self.conversation_phase = ConversationPhase.PAPER_ANALYSIS
        self.progress_status = "Analyzing paper quality..."

        # Grade all papers
        graded_papers = []
        for paper in self.recent_papers[:10]:  # Limit to 10 papers for performance
            score = self.grade_paper_quality(paper)
            self.paper_scores[paper.paper_id] = score
            graded_papers.append((paper, score))

        # Sort by overall score
        graded_papers.sort(key=lambda x: x[1].overall_score, reverse=True)

        response = f"### 📊 Paper Quality Assessment Results\n\n"
        response += f"**Analyzed {len(graded_papers)} papers** with comprehensive grading:\n\n"

        for i, (paper, score) in enumerate(graded_papers[:5], 1):
            grade_color = self._get_grade_color(score.overall_grade)
            stars = self._get_star_rating(score.overall_score)

            response += f'<div class="graded-paper-card">\n'
            response += f'<div class="paper-rank">#{i}</div>\n'
            response += f'<div class="paper-grade" style="color: {grade_color};">{score.overall_grade.value} ({score.overall_score}/10) {stars}</div>\n'
            response += f'<div class="paper-title">{paper.title}</div>\n'
            response += f'<div class="grade-breakdown">\n'
            response += f'• **Methodology:** {score.methodology_score}/10\n'
            response += f'• **Citation Impact:** {score.citation_impact_score}/10\n'
            response += f'• **Novelty:** {score.novelty_score}/10\n'
            response += f'• **Clarity:** {score.clarity_score}/10\n'
            response += f'• **Applicability:** {score.applicability_score}/10\n'
            response += f'</div>\n'
            response += f'<div class="grade-explanation">{score.explanation}</div>\n'

            if score.strengths:
                response += f'<div class="strengths">**Strengths:** {", ".join(score.strengths[:2])}</div>\n'

            response += f'</div>\n\n'

        if len(graded_papers) > 5:
            response += f'<div class="show-more-section">\n'
            response += f'<p><strong>📊 {len(graded_papers) - 5} more graded papers available</strong></p>\n'
            response += f'<p>💡 <em>Say "show more graded papers" or "show all graded papers" to see the complete assessment</em></p>\n'
            response += f'</div>\n\n'

        response += """**Next Steps:**
• Click 'Analyze Research Gaps' to identify opportunities
• Ask about specific papers or grading criteria
• Request papers from a different topic or time period"""

        return response

    def _get_grade_color(self, grade: PaperGrade) -> str:
        """Get color for grade display"""
        grade_colors = {
            PaperGrade.A_PLUS: "#00C851",
            PaperGrade.A: "#00C851",
            PaperGrade.A_MINUS: "#2BBBAD",
            PaperGrade.B_PLUS: "#4285F4",
            PaperGrade.B: "#4285F4",
            PaperGrade.B_MINUS: "#FF8800",
            PaperGrade.C_PLUS: "#FF8800",
            PaperGrade.C: "#FF6600",
            PaperGrade.C_MINUS: "#FF4444",
            PaperGrade.D: "#CC0000",
            PaperGrade.F: "#CC0000"
        }
        return grade_colors.get(grade, "#666666")

    def _get_star_rating(self, score: float) -> str:
        """Convert score to star rating"""
        full_stars = int(score // 2)
        half_star = 1 if (score % 2) >= 1 else 0
        empty_stars = 5 - full_stars - half_star

        return "⭐" * full_stars + "⭐" * half_star + "☆" * empty_stars

    def _handle_domain_refinement(self, user_input: str) -> str:
        """Handle domain refinement conversation"""
        # Extract interests from user input
        interests = self._extract_interests(user_input)
        self.user_interests.extend(interests)

        self.conversation_phase = ConversationPhase.LITERATURE_REVIEW

        return f"""### 🎯 Research Focus Refined

Great! I've noted your interests in: **{', '.join(interests)}**

**Now I'll search for papers** that match your refined focus. This will help me:

• Find the most relevant recent publications
• Identify specific research gaps in your area of interest
• Provide targeted recommendations based on your expertise level

**Searching academic databases...** 📚

Let me find recent papers in your area of interest."""

    def _extract_interests(self, user_input: str) -> List[str]:
        """Extract research interests from user input"""
        # Simple keyword extraction - could be enhanced with NLP
        interest_keywords = [
            'machine learning', 'deep learning', 'artificial intelligence', 'neural networks',
            'natural language processing', 'computer vision', 'robotics', 'data science',
            'healthcare', 'medical', 'biomedical', 'clinical', 'pharmaceutical',
            'energy', 'renewable', 'sustainable', 'environment', 'climate',
            'finance', 'fintech', 'blockchain', 'cryptocurrency',
            'education', 'e-learning', 'pedagogy', 'curriculum',
            'security', 'cybersecurity', 'privacy', 'cryptography'
        ]

        lower_input = user_input.lower()
        found_interests = [keyword for keyword in interest_keywords if keyword in lower_input]

        # If no specific keywords found, try to extract noun phrases (simplified)
        if not found_interests:
            words = user_input.split()
            # Look for potential research terms (2-3 word combinations)
            for i in range(len(words) - 1):
                if len(words[i]) > 3 and len(words[i+1]) > 3:
                    potential_interest = f"{words[i]} {words[i+1]}"
                    if potential_interest not in found_interests:
                        found_interests.append(potential_interest)

        return found_interests[:3]  # Limit to 3 interests

    def _handle_show_more_papers(self) -> str:
        """Handle request to show more papers"""
        if not self.recent_papers:
            return """### 📄 No Papers to Show

I don't have any papers loaded yet. Please search for papers first by:
• Telling me your research topic
• Clicking 'Find Papers' in the sidebar
• Saying "find papers on [your topic]"

**What research area would you like me to search?**"""

        if len(self.recent_papers) <= 5:
            return f"""### 📄 All Papers Already Shown

I found {len(self.recent_papers)} papers total, and they're all displayed above.

**Next steps:**
• **Analyze Research Gaps** - Identify opportunities in these papers
• **Search Different Topic** - Explore another research area
• **Refine Search** - Try more specific keywords

**What would you like to do next?**"""

        # Show all papers with grading
        response = f"### 📚 Complete Paper List ({len(self.recent_papers)} papers)\n\n"

        # Grade all papers if not already done
        graded_papers = []
        for paper in self.recent_papers:
            if paper.paper_id in self.paper_scores:
                score = self.paper_scores[paper.paper_id]
            else:
                score = self.grade_paper_quality(paper)
                self.paper_scores[paper.paper_id] = score
            graded_papers.append((paper, score))

        # Sort by quality
        graded_papers.sort(key=lambda x: x[1].overall_score, reverse=True)

        for i, (paper, score) in enumerate(graded_papers, 1):
            authors_str = ", ".join(paper.authors[:3])
            if len(paper.authors) > 3:
                authors_str += " et al."

            grade_color = self._get_grade_color(score.overall_grade)
            stars = self._get_star_rating(score.overall_score)

            response += f'<div class="enhanced-paper-card">\n'
            response += f'<div class="paper-header">\n'
            response += f'<span class="paper-rank">#{i}</span>\n'
            response += f'<span class="paper-grade" style="color: {grade_color};">{score.overall_grade.value} {stars}</span>\n'
            response += f'</div>\n'
            response += f'<div class="paper-title">{paper.title}</div>\n'
            response += f'<div class="paper-metadata">\n'
            response += f'<strong>Authors:</strong> {authors_str}<br>\n'
            response += f'<strong>Year:</strong> {paper.year} | <strong>Venue:</strong> {paper.venue}<br>\n'
            response += f'<strong>Citations:</strong> {paper.citations} | <strong>Quality Score:</strong> {score.overall_score}/10\n'
            if paper.doi:
                response += f'<br><strong>DOI:</strong> <a href="https://doi.org/{paper.doi}" target="_blank">{paper.doi}</a>\n'
            response += f'</div>\n'
            response += f'<div class="quality-summary">{score.explanation}</div>\n'
            response += f'</div>\n\n'

        response += f"""**📊 Summary Statistics:**
• **Total Papers:** {len(self.recent_papers)}
• **Average Quality:** {sum(score.overall_score for _, score in graded_papers) / len(graded_papers):.1f}/10
• **Top Grade:** {graded_papers[0][1].overall_grade.value}
• **Research Topic:** {self.current_topic}

**🎯 Next Steps:**
• **Analyze Research Gaps** - Find opportunities in this literature
• **Focus on Top Papers** - Dive deeper into A-grade research
• **Search Related Topics** - Explore connected research areas"""

        return response

    def _handle_show_more_gaps(self) -> str:
        """Handle request to show more research gaps"""
        if not self.identified_gaps:
            return """### 🔍 No Research Gaps Available

I haven't analyzed any research gaps yet. To get started:

1. **Search for papers** on your research topic
2. **Let me analyze** the literature for gaps
3. **Get comprehensive gap analysis** with recommendations

**Say:** "analyze research gaps" or click the 'Analyze Gaps' button

**What research topic should I analyze for gaps?**"""

        if len(self.identified_gaps) <= 4:
            return f"""### 🔍 All Research Gaps Already Shown

I identified {len(self.identified_gaps)} research gaps total, and they're all displayed above.

**To find more opportunities:**
• **Search broader topic** - Try related research areas
• **Search different time period** - Look at older or newer papers
• **Explore interdisciplinary connections** - Cross-domain research

**What would you like to explore next?**"""

        # Show all gaps with full details
        prioritized_gaps = self._prioritize_gaps_by_expertise(self.identified_gaps)

        response = f"### 🔍 Complete Research Gap Analysis ({len(prioritized_gaps)} opportunities)\n\n"
        response += f"**📊 Comprehensive analysis** for '{self.current_topic}'"

        if self.user_expertise:
            response += f" tailored to your **{self.user_expertise.value}** level"

        response += ":\n\n"

        gap_type_emojis = {
            "temporal": "📅",
            "methodological": "🔬",
            "theoretical": "💡",
            "application": "🚀",
            "interdisciplinary": "🔗",
            "technological": "⚡",
            "venue_diversity": "🌐",
            "geographical": "🗺️",
            "demographic": "👥"
        }

        for i, gap in enumerate(prioritized_gaps, 1):
            emoji = gap_type_emojis.get(gap.gap_type.lower(), "🔍")
            novelty_indicator = self._get_priority_indicator(gap.novelty_grade)
            impact_indicator = self._get_priority_indicator(gap.impact_grade)

            response += f'<div class="enhanced-gap-card">\n'
            response += f'<div class="gap-header">\n'
            response += f'<span class="gap-type">{emoji} {gap.gap_type.title()} Gap</span>\n'
            response += f'<span class="gap-priority">Novelty: {novelty_indicator} | Impact: {impact_indicator}</span>\n'
            response += f'</div>\n\n'

            response += f'**Gap {i}: {gap.description}**\n\n'

            response += f'**🎯 Why This Matters:**\n{gap.potential_impact}\n\n'

            if gap.supporting_evidence:
                response += f'**📋 Supporting Evidence:**\n'
                for evidence in gap.supporting_evidence:
                    response += f'• {evidence}\n'
                response += '\n'

            if gap.research_suggestions:
                response += f'**💡 Research Suggestions:**\n'
                for suggestion in gap.research_suggestions:
                    response += f'• {suggestion}\n'
                response += '\n'

            if gap.methodology_suggestions:
                response += f'**🔬 Methodology Ideas:**\n'
                for method in gap.methodology_suggestions:
                    response += f'• {method}\n'
                response += '\n'

            response += f'**⏱️ Estimated Timeline:** {gap.estimated_timeline}\n'
            response += f'**🎓 Required Expertise:** {gap.required_expertise.value.title()}\n'
            response += f'**📊 Confidence Score:** {gap.confidence_score:.2f}/1.0\n'

            if gap.required_resources:
                response += f'**📚 Resources Needed:** {", ".join(gap.required_resources)}\n'

            response += f'</div>\n\n'

        # Add summary statistics
        high_novelty = sum(1 for gap in prioritized_gaps if gap.novelty_grade == "High")
        high_impact = sum(1 for gap in prioritized_gaps if gap.impact_grade == "High")
        avg_confidence = sum(gap.confidence_score for gap in prioritized_gaps) / len(prioritized_gaps)

        response += f"""**📊 Gap Analysis Summary:**
• **Total Opportunities:** {len(prioritized_gaps)}
• **High Novelty Gaps:** {high_novelty}
• **High Impact Gaps:** {high_impact}
• **Average Confidence:** {avg_confidence:.2f}/1.0
• **Research Topic:** {self.current_topic}

**🚀 Recommended Next Steps:**
• **Pick a gap** that matches your expertise and interests
• **Develop a research proposal** using the methodology suggestions
• **Seek collaborators** in the identified resource areas
• **Start with pilot studies** to validate the research direction"""

        return response

    def _handle_show_more_graded_papers(self) -> str:
        """Handle request to show more graded papers"""
        if not self.paper_scores:
            return """### 📊 No Graded Papers Available

I haven't graded any papers yet. To get started:

1. **Search for papers** on your research topic
2. **Grade papers** using the "⭐ Grade Papers" button
3. **See comprehensive quality assessment** for all papers

**Say:** "grade papers" or search for papers first

**What research topic should I search and grade?**"""

        if len(self.paper_scores) <= 5:
            return f"""### 📊 All Graded Papers Already Shown

I have graded {len(self.paper_scores)} papers total, and they're all displayed above.

**To grade more papers:**
• **Search different topic** - Find papers in another research area
• **Search broader terms** - Get more papers to grade
• **Search different time period** - Look at older or newer papers

**What would you like to explore next?**"""

        # Show all graded papers with complete details
        graded_papers = []
        for paper_id, score in self.paper_scores.items():
            # Find the corresponding paper
            paper = next((p for p in self.recent_papers if p.paper_id == paper_id), None)
            if paper:
                graded_papers.append((paper, score))

        # Sort by quality score
        graded_papers.sort(key=lambda x: x[1].overall_score, reverse=True)

        response = f"### 📊 Complete Paper Quality Assessment ({len(graded_papers)} papers)\n\n"
        response += f"**✨ Comprehensive grading analysis** for all papers:\n\n"

        for i, (paper, score) in enumerate(graded_papers, 1):
            authors_str = ", ".join(paper.authors[:3])
            if len(paper.authors) > 3:
                authors_str += " et al."

            grade_color = self._get_grade_color(score.overall_grade)
            stars = self._get_star_rating(score.overall_score)

            response += f'<div class="graded-paper-card">\n'
            response += f'<div class="paper-header">\n'
            response += f'<span class="paper-rank">#{i}</span>\n'
            response += f'<span class="paper-grade" style="color: {grade_color};">{score.overall_grade.value} ({score.overall_score}/10) {stars}</span>\n'
            response += f'</div>\n'
            response += f'<div class="paper-title">{paper.title}</div>\n'
            response += f'<div class="paper-metadata">\n'
            response += f'<strong>Authors:</strong> {authors_str}<br>\n'
            response += f'<strong>Year:</strong> {paper.year} | <strong>Venue:</strong> {paper.venue}<br>\n'
            response += f'<strong>Citations:</strong> {paper.citations} | <strong>Quality Score:</strong> {score.overall_score}/10\n'
            if paper.doi:
                response += f'<br><strong>DOI:</strong> <a href="https://doi.org/{paper.doi}" target="_blank">{paper.doi}</a>\n'
            response += f'</div>\n'

            # Detailed grade breakdown
            response += f'<div class="grade-breakdown">\n'
            response += f'**📊 Detailed Quality Breakdown:**\n'
            response += f'• **Methodology Rigor:** {score.methodology_score}/10\n'
            response += f'• **Citation Impact:** {score.citation_impact_score}/10\n'
            response += f'• **Novelty & Innovation:** {score.novelty_score}/10\n'
            response += f'• **Clarity of Presentation:** {score.clarity_score}/10\n'
            response += f'• **Practical Applicability:** {score.applicability_score}/10\n'
            response += f'</div>\n'

            response += f'<div class="grade-explanation">**Assessment:** {score.explanation}</div>\n'

            if score.strengths:
                response += f'<div class="strengths">**✅ Strengths:** {", ".join(score.strengths)}</div>\n'

            if score.weaknesses:
                response += f'<div class="weaknesses">**⚠️ Areas for Improvement:** {", ".join(score.weaknesses)}</div>\n'

            response += f'</div>\n\n'

        # Add comprehensive statistics
        avg_score = sum(score.overall_score for _, score in graded_papers) / len(graded_papers)
        grade_distribution = {}
        for _, score in graded_papers:
            grade = score.overall_grade.value
            grade_distribution[grade] = grade_distribution.get(grade, 0) + 1

        response += f"""**📈 Quality Assessment Summary:**
• **Total Papers Graded:** {len(graded_papers)}
• **Average Quality Score:** {avg_score:.1f}/10
• **Highest Grade:** {graded_papers[0][1].overall_grade.value} ({graded_papers[0][1].overall_score}/10)
• **Research Topic:** {self.current_topic}

**📊 Grade Distribution:**"""

        for grade, count in sorted(grade_distribution.items()):
            response += f"\n• **{grade} Grade:** {count} papers"

        response += f"""

**🎯 Next Steps:**
• **Analyze Research Gaps** - Find opportunities in this high-quality literature
• **Focus on Top Papers** - Deep dive into A+ and A grade research
• **Export Data** - Save this analysis for your research
• **Search Related Topics** - Explore connected research areas"""

        return response

    def export_data_as_json(self) -> str:
        """Export research data as JSON"""
        export_data = {
            "metadata": {
                "export_date": datetime.now().isoformat(),
                "topic": self.current_topic,
                "user_expertise": self.user_expertise.value if self.user_expertise else None,
                "total_papers": len(self.recent_papers),
                "total_gaps": len(self.identified_gaps),
                "graded_papers": len(self.paper_scores)
            },
            "papers": [],
            "paper_grades": {},
            "research_gaps": []
        }

        # Export papers
        for paper in self.recent_papers:
            paper_data = {
                "title": paper.title,
                "authors": paper.authors,
                "year": paper.year,
                "venue": paper.venue,
                "abstract": paper.abstract,
                "citations": paper.citations,
                "doi": paper.doi,
                "url": paper.url,
                "source": paper.source
            }
            export_data["papers"].append(paper_data)

        # Export paper grades
        for paper_id, score in self.paper_scores.items():
            paper = next((p for p in self.recent_papers if p.paper_id == paper_id), None)
            if paper:
                export_data["paper_grades"][paper.title] = {
                    "overall_grade": score.overall_grade.value,
                    "overall_score": score.overall_score,
                    "methodology_score": score.methodology_score,
                    "citation_impact_score": score.citation_impact_score,
                    "novelty_score": score.novelty_score,
                    "clarity_score": score.clarity_score,
                    "applicability_score": score.applicability_score,
                    "explanation": score.explanation,
                    "strengths": score.strengths,
                    "weaknesses": score.weaknesses
                }

        # Export research gaps
        for gap in self.identified_gaps:
            gap_data = {
                "gap_type": gap.gap_type,
                "description": gap.description,
                "supporting_evidence": gap.supporting_evidence,
                "potential_impact": gap.potential_impact,
                "suggested_approach": gap.suggested_approach,
                "confidence_score": gap.confidence_score,
                "novelty_grade": gap.novelty_grade,
                "impact_grade": gap.impact_grade,
                "research_suggestions": gap.research_suggestions,
                "methodology_suggestions": gap.methodology_suggestions,
                "required_expertise": gap.required_expertise.value,
                "estimated_timeline": gap.estimated_timeline,
                "required_resources": gap.required_resources
            }
            export_data["research_gaps"].append(gap_data)

        return json.dumps(export_data, indent=2, ensure_ascii=False)

    def export_data_as_yaml(self) -> str:
        """Export research data as YAML"""
        # Get JSON data and convert to YAML
        json_data = json.loads(self.export_data_as_json())
        return yaml.dump(json_data, default_flow_style=False, allow_unicode=True, sort_keys=False)

    def export_data_as_csv(self) -> Tuple[str, str, str]:
        """Export research data as multiple CSV files"""

        # Papers CSV
        papers_data = []
        for paper in self.recent_papers:
            # Get grade if available
            grade_info = {}
            if paper.paper_id in self.paper_scores:
                score = self.paper_scores[paper.paper_id]
                grade_info = {
                    "Overall_Grade": score.overall_grade.value,
                    "Overall_Score": score.overall_score,
                    "Methodology_Score": score.methodology_score,
                    "Citation_Impact_Score": score.citation_impact_score,
                    "Novelty_Score": score.novelty_score,
                    "Clarity_Score": score.clarity_score,
                    "Applicability_Score": score.applicability_score,
                    "Explanation": score.explanation,
                    "Strengths": "; ".join(score.strengths),
                    "Weaknesses": "; ".join(score.weaknesses)
                }

            paper_row = {
                "Title": paper.title,
                "Authors": "; ".join(paper.authors),
                "Year": paper.year,
                "Venue": paper.venue,
                "Abstract": paper.abstract,
                "Citations": paper.citations,
                "DOI": paper.doi,
                "URL": paper.url,
                "Source": paper.source,
                **grade_info
            }
            papers_data.append(paper_row)

        papers_df = pd.DataFrame(papers_data)
        papers_csv = papers_df.to_csv(index=False)

        # Research Gaps CSV
        gaps_data = []
        for gap in self.identified_gaps:
            gap_row = {
                "Gap_Type": gap.gap_type,
                "Description": gap.description,
                "Supporting_Evidence": "; ".join(gap.supporting_evidence),
                "Potential_Impact": gap.potential_impact,
                "Suggested_Approach": gap.suggested_approach,
                "Confidence_Score": gap.confidence_score,
                "Novelty_Grade": gap.novelty_grade,
                "Impact_Grade": gap.impact_grade,
                "Research_Suggestions": "; ".join(gap.research_suggestions),
                "Methodology_Suggestions": "; ".join(gap.methodology_suggestions),
                "Required_Expertise": gap.required_expertise.value,
                "Estimated_Timeline": gap.estimated_timeline,
                "Required_Resources": "; ".join(gap.required_resources)
            }
            gaps_data.append(gap_row)

        gaps_df = pd.DataFrame(gaps_data)
        gaps_csv = gaps_df.to_csv(index=False)

        # Summary CSV
        summary_data = [{
            "Export_Date": datetime.now().isoformat(),
            "Research_Topic": self.current_topic,
            "User_Expertise": self.user_expertise.value if self.user_expertise else "Not Set",
            "Total_Papers": len(self.recent_papers),
            "Graded_Papers": len(self.paper_scores),
            "Research_Gaps_Identified": len(self.identified_gaps),
            "Average_Paper_Quality": sum(score.overall_score for score in self.paper_scores.values()) / len(self.paper_scores) if self.paper_scores else 0
        }]

        summary_df = pd.DataFrame(summary_data)
        summary_csv = summary_df.to_csv(index=False)

        return papers_csv, gaps_csv, summary_csv

    def export_data_as_excel(self) -> bytes:
        """Export research data as Excel file with multiple sheets"""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel packages not available. Please install openpyxl and xlsxwriter.")

        # Create Excel file in memory
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Papers sheet
            papers_data = []
            for paper in self.recent_papers:
                # Get grade if available
                grade_info = {}
                if paper.paper_id in self.paper_scores:
                    score = self.paper_scores[paper.paper_id]
                    grade_info = {
                        "Overall_Grade": score.overall_grade.value,
                        "Overall_Score": score.overall_score,
                        "Methodology_Score": score.methodology_score,
                        "Citation_Impact_Score": score.citation_impact_score,
                        "Novelty_Score": score.novelty_score,
                        "Clarity_Score": score.clarity_score,
                        "Applicability_Score": score.applicability_score,
                        "Explanation": score.explanation,
                        "Strengths": "; ".join(score.strengths),
                        "Weaknesses": "; ".join(score.weaknesses)
                    }

                paper_row = {
                    "Title": paper.title,
                    "Authors": "; ".join(paper.authors),
                    "Year": paper.year,
                    "Venue": paper.venue,
                    "Abstract": paper.abstract[:500] + "..." if len(paper.abstract) > 500 else paper.abstract,  # Truncate for Excel
                    "Citations": paper.citations,
                    "DOI": paper.doi,
                    "URL": paper.url,
                    "Source": paper.source,
                    **grade_info
                }
                papers_data.append(paper_row)

            papers_df = pd.DataFrame(papers_data)
            papers_df.to_excel(writer, sheet_name='Papers', index=False)

            # Research Gaps sheet
            gaps_data = []
            for gap in self.identified_gaps:
                gap_row = {
                    "Gap_Type": gap.gap_type,
                    "Description": gap.description,
                    "Supporting_Evidence": "; ".join(gap.supporting_evidence),
                    "Potential_Impact": gap.potential_impact,
                    "Suggested_Approach": gap.suggested_approach,
                    "Confidence_Score": gap.confidence_score,
                    "Novelty_Grade": gap.novelty_grade,
                    "Impact_Grade": gap.impact_grade,
                    "Research_Suggestions": "; ".join(gap.research_suggestions),
                    "Methodology_Suggestions": "; ".join(gap.methodology_suggestions),
                    "Required_Expertise": gap.required_expertise.value,
                    "Estimated_Timeline": gap.estimated_timeline,
                    "Required_Resources": "; ".join(gap.required_resources)
                }
                gaps_data.append(gap_row)

            if gaps_data:
                gaps_df = pd.DataFrame(gaps_data)
                gaps_df.to_excel(writer, sheet_name='Research_Gaps', index=False)

            # Summary sheet
            summary_data = [{
                "Export_Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Research_Topic": self.current_topic,
                "User_Expertise": self.user_expertise.value if self.user_expertise else "Not Set",
                "Total_Papers": len(self.recent_papers),
                "Graded_Papers": len(self.paper_scores),
                "Research_Gaps_Identified": len(self.identified_gaps),
                "Average_Paper_Quality": round(sum(score.overall_score for score in self.paper_scores.values()) / len(self.paper_scores), 2) if self.paper_scores else 0
            }]

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Format the Excel file
            workbook = writer.book

            # Format Papers sheet
            if 'Papers' in workbook.sheetnames:
                worksheet = workbook['Papers']

                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return output.getvalue()
    
    def _handle_paper_search(self, user_input: str) -> str:
        """Enhanced paper search with automatic quality grading"""
        topic = self.current_topic or self.extract_research_topic(user_input)

        if not topic:
            return """### 🔍 Let's Find Research Papers!

I need a research topic to search for papers. Here are some examples:

**Popular Research Areas:**
• "machine learning in healthcare"
• "sustainable energy storage"
• "natural language processing"
• "computer vision applications"
• "blockchain technology"

**What research area interests you?**"""

        self.current_topic = topic
        self.conversation_phase = ConversationPhase.LITERATURE_REVIEW
        self.progress_status = f"Searching for papers on '{topic}'..."

        # Search for papers with status updates
        def status_update(message):
            # In a real implementation, this could update a progress bar
            # For now, we'll log the status
            logger.info(f"Search Status: {message}")

        papers = self.find_recent_papers(topic, status_update)
        self.recent_papers = papers

        if not papers:
            return f"""### ❌ No Papers Found for '{topic}'

This might be due to:
• **Search term too specific** - try broader terms
• **Network connectivity issues** - check your connection
• **API rate limits** - try again in a few minutes

**Suggestions:**
• Try a broader search term (e.g., "machine learning" instead of "deep reinforcement learning for autonomous vehicles")
• Check spelling and try alternative terms
• Consider related research areas

**What would you like to try next?**"""

        # Automatically grade papers for enhanced experience
        self.progress_status = f"Analyzing quality of {len(papers)} papers..."

        # Grade top papers
        graded_papers = []
        for paper in papers[:8]:  # Grade top 8 papers
            score = self.grade_paper_quality(paper)
            self.paper_scores[paper.paper_id] = score
            graded_papers.append((paper, score))

        # Sort by quality score
        graded_papers.sort(key=lambda x: x[1].overall_score, reverse=True)

        # Format response with enhanced paper cards including grades
        response = f"### 📚 Found {len(papers)} Recent Papers on '{topic}'\n\n"
        response += f"**✨ Automatically graded for quality** - showing top {min(5, len(graded_papers))} papers:\n\n"

        for i, (paper, score) in enumerate(graded_papers[:5], 1):
            authors_str = ", ".join(paper.authors[:3])
            if len(paper.authors) > 3:
                authors_str += " et al."

            grade_color = self._get_grade_color(score.overall_grade)
            stars = self._get_star_rating(score.overall_score)

            response += f'<div class="enhanced-paper-card">\n'
            response += f'<div class="paper-header">\n'
            response += f'<span class="paper-rank">#{i}</span>\n'
            response += f'<span class="paper-grade" style="color: {grade_color};">{score.overall_grade.value} {stars}</span>\n'
            response += f'</div>\n'
            response += f'<div class="paper-title">{paper.title}</div>\n'
            response += f'<div class="paper-metadata">\n'
            response += f'<strong>Authors:</strong> {authors_str}<br>\n'
            response += f'<strong>Year:</strong> {paper.year} | <strong>Venue:</strong> {paper.venue}<br>\n'
            response += f'<strong>Citations:</strong> {paper.citations} | <strong>Quality Score:</strong> {score.overall_score}/10\n'
            if paper.doi:
                response += f'<br><strong>DOI:</strong> <a href="https://doi.org/{paper.doi}" target="_blank">{paper.doi}</a>\n'
            response += f'</div>\n'
            response += f'<div class="quality-summary">{score.explanation}</div>\n'
            response += f'</div>\n\n'

        if len(papers) > 5:
            response += f'<div class="show-more-section">\n'
            response += f'<p><strong>📄 {len(papers) - 5} more papers available</strong></p>\n'
            response += f'<p>💡 <em>Say "show more papers" or "show all papers" to see the complete list</em></p>\n'
            response += f'</div>\n\n'

        response += f"""**🎯 Next Steps:**
• **Analyze Research Gaps** - I'll identify opportunities based on these papers
• **View Detailed Grades** - See comprehensive quality assessment for all papers
• **Refine Search** - Search for papers in a specific sub-area
• **Change Topic** - Explore a different research area

**What would you like to do next?**"""

        return response
    
    def _handle_gap_analysis(self) -> str:
        """Enhanced gap analysis with personalized recommendations"""
        if not self.recent_papers:
            return """### 🔍 Research Gap Analysis

I need to search for papers first before analyzing gaps.

**To get started:**
1. Tell me your research topic
2. I'll search for recent papers
3. Then I'll identify specific research opportunities

**What research area should I analyze for gaps?**"""

        self.conversation_phase = ConversationPhase.GAP_ANALYSIS
        self.progress_status = f"Analyzing research gaps in {len(self.recent_papers)} papers..."

        # Enhanced gap analysis with user expertise consideration
        gaps = self.analyze_research_gaps(self.recent_papers)
        self.identified_gaps = gaps

        if not gaps:
            return f"""### 🔍 Gap Analysis Complete

I analyzed {len(self.recent_papers)} papers but couldn't identify clear research gaps.

**This might mean:**
• The field is well-researched with few obvious gaps
• The search was too narrow - try broader terms
• More recent papers are needed

**Suggestions:**
• Search for papers in a related sub-field
• Try a different time period (e.g., last 2 years)
• Explore interdisciplinary connections

**What would you like to try next?**"""

        # Filter and prioritize gaps based on user expertise
        prioritized_gaps = self._prioritize_gaps_by_expertise(gaps)

        response = f"### 🔍 Research Gap Analysis for '{self.current_topic}'\n\n"
        response += f"**📊 Analysis Complete!** Based on {len(self.recent_papers)} recent papers, I've identified **{len(gaps)} research opportunities**"

        if self.user_expertise:
            response += f" tailored to your **{self.user_expertise.value}** level"

        response += ":\n\n"

        gap_type_emojis = {
            "temporal": "📅",
            "methodological": "🔬",
            "theoretical": "💡",
            "application": "🚀",
            "interdisciplinary": "🔗",
            "technological": "⚡"
        }

        for i, gap in enumerate(prioritized_gaps[:4], 1):  # Show top 4 gaps
            emoji = gap_type_emojis.get(gap.gap_type.lower(), "🔍")

            # Get priority indicators
            novelty_indicator = self._get_priority_indicator(gap.novelty_grade)
            impact_indicator = self._get_priority_indicator(gap.impact_grade)

            response += f'<div class="enhanced-gap-card">\n'
            response += f'<div class="gap-header">\n'
            response += f'<span class="gap-type">{emoji} {gap.gap_type.title()} Gap</span>\n'
            response += f'<span class="gap-priority">Novelty: {novelty_indicator} | Impact: {impact_indicator}</span>\n'
            response += f'</div>\n\n'

            response += f'**Gap {i}: {gap.description}**\n\n'

            response += f'**🎯 Why This Matters:**\n{gap.potential_impact}\n\n'

            response += f'**📋 Research Suggestions:**\n'
            for suggestion in gap.research_suggestions[:2]:
                response += f'• {suggestion}\n'

            if gap.methodology_suggestions:
                response += f'\n**🔬 Methodology Ideas:**\n'
                for method in gap.methodology_suggestions[:2]:
                    response += f'• {method}\n'

            response += f'\n**⏱️ Estimated Timeline:** {gap.estimated_timeline}\n'
            response += f'**🎓 Required Expertise:** {gap.required_expertise.value.title()}\n'

            if gap.required_resources:
                response += f'**📚 Resources Needed:** {", ".join(gap.required_resources[:3])}\n'

            response += f'</div>\n\n'

        if len(gaps) > 4:
            response += f'<div class="show-more-section">\n'
            response += f'<p><strong>🔍 {len(gaps) - 4} more research opportunities identified</strong></p>\n'
            response += f'<p>💡 <em>Say "show more gaps" or "show all gaps" to see the complete analysis</em></p>\n'
            response += f'</div>\n\n'

        self.conversation_phase = ConversationPhase.RECOMMENDATION_REVIEW

        response += f"""**🚀 Next Steps:**
• **Get Detailed Recommendations** - Deep dive into specific gaps
• **Find Collaborators** - I can suggest potential research partners
• **Explore Methodologies** - Get specific research approaches
• **Search Related Areas** - Find connected research opportunities

**Which gap interests you most, or what would you like to explore next?**"""

        return response

    def _prioritize_gaps_by_expertise(self, gaps: List[ResearchGap]) -> List[ResearchGap]:
        """Prioritize research gaps based on user expertise level"""
        if not self.user_expertise:
            return gaps  # Return as-is if no expertise level set

        # Filter gaps appropriate for user's expertise level
        suitable_gaps = []
        for gap in gaps:
            if self._is_gap_suitable_for_expertise(gap, self.user_expertise):
                suitable_gaps.append(gap)

        # Sort by novelty and impact grades
        def gap_priority_score(gap):
            novelty_score = {"High": 3, "Medium": 2, "Low": 1}.get(gap.novelty_grade, 2)
            impact_score = {"High": 3, "Medium": 2, "Low": 1}.get(gap.impact_grade, 2)
            confidence_score = gap.confidence_score
            return (novelty_score + impact_score) * confidence_score

        suitable_gaps.sort(key=gap_priority_score, reverse=True)
        return suitable_gaps

    def _is_gap_suitable_for_expertise(self, gap: ResearchGap, user_expertise: UserExpertiseLevel) -> bool:
        """Check if a research gap is suitable for user's expertise level"""
        expertise_hierarchy = {
            UserExpertiseLevel.BEGINNER: 1,
            UserExpertiseLevel.INTERMEDIATE: 2,
            UserExpertiseLevel.ADVANCED: 3,
            UserExpertiseLevel.EXPERT: 4
        }

        gap_requirement = expertise_hierarchy.get(gap.required_expertise, 2)
        user_level = expertise_hierarchy.get(user_expertise, 2)

        # Allow gaps that are at or slightly above user's level
        return gap_requirement <= user_level + 1

    def _get_priority_indicator(self, grade: str) -> str:
        """Get visual indicator for priority grades"""
        indicators = {
            "High": "🔥 High",
            "Medium": "⚡ Medium",
            "Low": "💡 Low"
        }
        return indicators.get(grade, "⚡ Medium")
    
    def _handle_topic_introduction(self, topic: str) -> str:
        """Enhanced topic introduction with guided next steps"""
        self.conversation_phase = ConversationPhase.DOMAIN_REFINEMENT

        response = f"### 🎯 Excellent Choice: '{topic}'\n\n"
        response += "I'm excited to help you explore this research area! Let me guide you through the process.\n\n"

        # Check if user expertise is set
        if not self.user_expertise:
            response += """### 🎓 **First, let's personalize your experience**

To provide the best recommendations, I'd like to know your expertise level:

• **Beginner** - New to this research area, need foundational guidance
• **Intermediate** - Some background knowledge, ready for moderate challenges
• **Advanced** - Strong foundation, interested in cutting-edge research
• **Expert** - Deep expertise, looking for novel opportunities

**Just say:** *"I'm a beginner"* or *"I'm an expert in [topic]"*

---

"""

        response += f"""### 🗺️ **Your Research Journey for '{topic}'**

**Next Steps I'll Guide You Through:**

**Step 1: 🔍 Paper Discovery & Grading**
• Search top academic databases (Semantic Scholar, Crossref)
• Find the most recent and relevant papers
• **Automatically grade each paper** (A+ to F) based on:
  - Methodology rigor
  - Citation impact
  - Novelty and innovation
  - Clarity of presentation
  - Practical applicability

**Step 2: 📊 Intelligent Gap Analysis**
• Analyze patterns across all papers
• Identify **specific research opportunities**
• Grade gaps by novelty (🔥 High, ⚡ Medium, 💡 Low)
• Assess potential impact and feasibility

**Step 3: 🎯 Personalized Recommendations**
• Suggest research directions matched to your expertise
• Provide concrete methodology approaches
• Estimate timelines and required resources
• Recommend potential collaborators

---

### 🚀 **Ready to Begin?**

**Option 1:** Start the paper search immediately
*Say: "Find recent papers" or click the button*

**Option 2:** Refine your focus first
*Tell me: "I'm specifically interested in [sub-area]"*

**Option 3:** Set your expertise level first
*Say: "I'm a [beginner/intermediate/advanced/expert]"*

**What would you like to do first?**"""

        return response
    
    def _handle_general_query(self, user_input: str) -> str:
        """Handle general queries"""
        return """I'm your Research Gap Finder assistant. I can help you:

• **Find recent academic papers** on any topic
• **Identify research gaps** and opportunities
• **Analyze publication trends**
• **Suggest research directions**

Just tell me what research area interests you, or click one of the quick action buttons to get started!

Example: "I want to research machine learning in healthcare" or "Find papers on renewable energy"
"""
    
    def start_conversation(self) -> str:
        """Enhanced welcome message with step-by-step guidance"""
        # Check what's available
        has_ai = bool(self.llm.providers)
        ai_providers = list(self.llm.providers.keys()) if has_ai else []

        welcome = """### 👋 Welcome to Research Gap Finder!

I'm your **AI-powered research assistant** designed to guide you through discovering research opportunities step-by-step. Let me help you navigate the world of academic research!"""

        if has_ai:
            provider_names = {
                "openai": "OpenAI GPT-4",
                "anthropic": "Anthropic Claude",
                "gemini": "Google Gemini"
            }
            active_names = [provider_names.get(p, p) for p in ai_providers]

            welcome += f"""

### 🚀 **Enhanced AI Mode Active**
*Powered by {', '.join(active_names)}*

**My Advanced Capabilities:**
• 📊 **Comprehensive Paper Grading** (A+ to F scale with detailed criteria)
• 🎯 **Personalized Gap Analysis** (tailored to your expertise level)
• 🔍 **Smart Research Recommendations** (with methodology suggestions)
• 📈 **Impact Assessment** (novelty and potential impact scoring)
• 🤝 **Collaboration Suggestions** (potential partners and resources)"""
        else:
            welcome += """

### 📚 **Basic Mode Active**
*No AI providers configured - but still powerful!*

**Available Features:**
• 🔍 Academic database search (Semantic Scholar & Crossref)
• 📄 Recent paper discovery (last 5 years)
• ✅ Paper validation and quality checks
• 📊 Basic gap analysis

💡 *Add an API key for OpenAI, Anthropic, or Google to unlock advanced AI features*"""

        welcome += """

---

### 🗺️ **Your Research Journey - Step by Step**

**Step 1: Tell Me About You** *(Optional but recommended)*
• What's your expertise level? (beginner/intermediate/advanced/expert)
• This helps me tailor recommendations to your capabilities

**Step 2: Share Your Research Interest**
Just tell me what fascinates you! Examples:
• *"I'm interested in machine learning for healthcare"*
• *"sustainable energy storage solutions"*
• *"social media impact on mental health"*

**Step 3: I'll Search & Grade Papers**
• 🔍 Search multiple academic databases
• 📊 Automatically grade papers for quality (A+ to F)
• ⭐ Show you the highest-quality research first

**Step 4: Discover Research Gaps**
• 🎯 Identify specific opportunities in your field
• 📈 Grade gaps by novelty and potential impact
• 🔬 Suggest concrete research approaches

**Step 5: Get Actionable Recommendations**
• 📋 Specific research suggestions tailored to your level
• ⏱️ Timeline estimates for research projects
• 🤝 Collaboration and resource recommendations

---

### 🚀 **Ready to Start?**

**Option 1:** Tell me your expertise level first
*"I'm a beginner in machine learning"* or *"I'm an expert in renewable energy"*

**Option 2:** Jump right into your research interest
*"Find papers on quantum computing applications"*

**Option 3:** Use the quick action buttons in the sidebar

**What would you like to explore today?**"""

        return welcome

# Session State Management
def initialize_session_state():
    """Initialize session state variables"""
    if 'research_assistant' not in st.session_state:
        st.session_state.research_assistant = ResearchGapAnalyzer()
    
    if 'messages' not in st.session_state:
        st.session_state.messages = []
        welcome = st.session_state.research_assistant.start_conversation()
        st.session_state.messages.append({
            'role': 'assistant',
            'content': welcome,
            'timestamp': datetime.now()
        })
    
    if 'thinking' not in st.session_state:
        st.session_state.thinking = False

# UI Functions
def display_header():
    """Display the main application header"""
    st.markdown("""
    <div class="main-header">
        <h1>🔬 Research Gap Finder</h1>
        <p>AI-powered tool for discovering research opportunities</p>
    </div>
    """, unsafe_allow_html=True)


def display_message(message):
    """Display a single chat message with enhanced formatting"""
    role_class = "user-message" if message['role'] == 'user' else "assistant-message"
    
    # Format timestamp
    timestamp = message.get('timestamp', datetime.now())
    if isinstance(timestamp, str):
        try:
            timestamp = datetime.fromisoformat(timestamp)
        except:
            timestamp = datetime.now()
    time_str = timestamp.strftime("%I:%M %p")
    
    # Get content
    content = message['content']
    
    # For assistant messages, use markdown directly
    if message['role'] == 'assistant':
        message_html = f"""
        <div class="chat-message {role_class}">
            <div class="message-content">
                {content}
            </div>
            <div class="message-timestamp">{time_str}</div>
        </div>
        """
        st.markdown(message_html, unsafe_allow_html=True)
    else:
        # For user messages, escape HTML
        safe_content = content.replace('<', '&lt;').replace('>', '&gt;')
        message_html = f"""
        <div class="chat-message {role_class}">
            <div class="message-content">{safe_content}</div>
            <div class="message-timestamp">{time_str}</div>
        </div>
        """
        st.markdown(message_html, unsafe_allow_html=True)


def display_thinking_indicator():
    """Display thinking animation"""
    st.markdown("""
    <div class="thinking-indicator">
        <div class="thinking-dots">
            <div class="thinking-dot"></div>
            <div class="thinking-dot"></div>
            <div class="thinking-dot"></div>
        </div>
        <span>Searching academic databases...</span>
    </div>
    """, unsafe_allow_html=True)


def send_message(message: str):
    """Process and send message with rate limiting status updates"""
    if not message or not message.strip():
        return

    # Add user message
    st.session_state.messages.append({
        'role': 'user',
        'content': message.strip(),
        'timestamp': datetime.now()
    })

    # Set thinking state
    st.session_state.thinking = True

    # Create status container for rate limiting feedback
    status_container = st.empty()

    def update_status(status_message):
        """Update status display for user feedback"""
        with status_container.container():
            if "rate limit" in status_message.lower():
                st.info(f"🕐 {status_message}")
            elif "searching" in status_message.lower():
                st.info(f"🔍 {status_message}")
            elif "waiting" in status_message.lower():
                st.warning(f"⏳ {status_message}")
            else:
                st.info(f"📊 {status_message}")

    # Process message
    try:
        start_time = time.time()

        # Set up status callback for the research assistant
        original_find_papers = st.session_state.research_assistant.find_recent_papers

        def find_papers_with_status(topic, status_callback=None):
            if status_callback is None:
                status_callback = update_status
            return original_find_papers(topic, status_callback)

        # Temporarily replace the method
        st.session_state.research_assistant.find_recent_papers = find_papers_with_status

        response = st.session_state.research_assistant.process_message(message)
        response_time = time.time() - start_time

        # Restore original method
        st.session_state.research_assistant.find_recent_papers = original_find_papers

        # Clear status
        status_container.empty()

        # Add assistant response
        st.session_state.messages.append({
            'role': 'assistant',
            'content': response,
            'timestamp': datetime.now(),
            'response_time': response_time
        })

    except Exception as e:
        # Clear status on error
        status_container.empty()

        error_message = f"I encountered an issue: {str(e)}\n\nPlease try again or rephrase your request."
        st.session_state.messages.append({
            'role': 'assistant',
            'content': error_message,
            'timestamp': datetime.now(),
            'error': True
        })

    finally:
        st.session_state.thinking = False
        st.rerun()


def display_sidebar():
    """Display sidebar with controls"""
    with st.sidebar:
        st.markdown("## 🎯 Research Assistant")
        
        # Quick Actions
        st.markdown("### Quick Actions")

        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔍 Find Papers", use_container_width=True):
                if st.session_state.research_assistant.current_topic:
                    send_message(f"find papers on {st.session_state.research_assistant.current_topic}")
                else:
                    st.warning("Set a topic first!")
        
        with col2:
            if st.button("📊 Analyze Gaps", use_container_width=True):
                if st.session_state.research_assistant.recent_papers:
                    send_message("analyze research gaps")
                else:
                    st.warning("Search papers first!")

        # Paper Grading Action
        if st.session_state.research_assistant.recent_papers:
            if st.button("⭐ Grade Papers", use_container_width=True):
                send_message("grade papers")

        # Show More Actions (if data available)
        if st.session_state.research_assistant.recent_papers or st.session_state.research_assistant.identified_gaps:
            st.markdown("### 📋 View More")

            col3, col4 = st.columns(2)

            with col3:
                if st.session_state.research_assistant.recent_papers and len(st.session_state.research_assistant.recent_papers) > 5:
                    if st.button("📄 Show All Papers", use_container_width=True):
                        send_message("show all papers")
                elif st.session_state.research_assistant.recent_papers:
                    st.info(f"All {len(st.session_state.research_assistant.recent_papers)} papers shown")

            with col4:
                if st.session_state.research_assistant.identified_gaps and len(st.session_state.research_assistant.identified_gaps) > 4:
                    if st.button("🔍 Show All Gaps", use_container_width=True):
                        send_message("show all gaps")
                elif st.session_state.research_assistant.identified_gaps:
                    st.info(f"All {len(st.session_state.research_assistant.identified_gaps)} gaps shown")

        # Graded Papers Section
        if st.session_state.research_assistant.paper_scores:
            st.markdown("### ⭐ Graded Papers")

            col5, col6 = st.columns(2)

            with col5:
                if len(st.session_state.research_assistant.paper_scores) > 5:
                    if st.button("📊 Show All Graded Papers", use_container_width=True):
                        send_message("show all graded papers")
                else:
                    st.info(f"All {len(st.session_state.research_assistant.paper_scores)} graded papers shown")

            with col6:
                if st.button("⭐ Re-grade Papers", use_container_width=True):
                    send_message("grade papers")

        # Export Data Section
        if st.session_state.research_assistant.recent_papers or st.session_state.research_assistant.identified_gaps:
            st.markdown("### 📊 Export Data")

            # Initialize export state
            if 'export_ready' not in st.session_state:
                st.session_state.export_ready = {}

            # JSON Export
            if st.button("📄 Export JSON", use_container_width=True, key="export_json_btn"):
                json_data = st.session_state.research_assistant.export_data_as_json()
                st.session_state.export_ready['json'] = {
                    'data': json_data,
                    'filename': f"research_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                }

            if 'json' in st.session_state.export_ready:
                st.download_button(
                    label="💾 Download JSON",
                    data=st.session_state.export_ready['json']['data'],
                    file_name=st.session_state.export_ready['json']['filename'],
                    mime="application/json",
                    use_container_width=True,
                    key="download_json_btn"
                )

            # YAML Export
            if st.button("📋 Export YAML", use_container_width=True, key="export_yaml_btn"):
                yaml_data = st.session_state.research_assistant.export_data_as_yaml()
                st.session_state.export_ready['yaml'] = {
                    'data': yaml_data,
                    'filename': f"research_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.yaml"
                }

            if 'yaml' in st.session_state.export_ready:
                st.download_button(
                    label="💾 Download YAML",
                    data=st.session_state.export_ready['yaml']['data'],
                    file_name=st.session_state.export_ready['yaml']['filename'],
                    mime="text/yaml",
                    use_container_width=True,
                    key="download_yaml_btn"
                )

            # CSV Export
            if st.button("📊 Export CSV", use_container_width=True, key="export_csv_btn"):
                papers_csv, gaps_csv, summary_csv = st.session_state.research_assistant.export_data_as_csv()

                # Create a zip file with all CSV files
                zip_buffer = io.BytesIO()

                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    zip_file.writestr("papers.csv", papers_csv)
                    zip_file.writestr("research_gaps.csv", gaps_csv)
                    zip_file.writestr("summary.csv", summary_csv)

                zip_buffer.seek(0)

                st.session_state.export_ready['csv'] = {
                    'data': zip_buffer.getvalue(),
                    'filename': f"research_data_csv_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                }

            if 'csv' in st.session_state.export_ready:
                st.download_button(
                    label="💾 Download CSV Bundle",
                    data=st.session_state.export_ready['csv']['data'],
                    file_name=st.session_state.export_ready['csv']['filename'],
                    mime="application/zip",
                    use_container_width=True,
                    key="download_csv_btn"
                )

            # Excel Export
            if EXCEL_AVAILABLE:
                if st.button("📈 Export Excel", use_container_width=True, key="export_excel_btn"):
                    try:
                        excel_data = st.session_state.research_assistant.export_data_as_excel()
                        st.session_state.export_ready['excel'] = {
                            'data': excel_data,
                            'filename': f"research_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        }
                    except Exception as e:
                        st.error(f"Excel export error: {e}")

                if 'excel' in st.session_state.export_ready:
                    st.download_button(
                        label="💾 Download Excel",
                        data=st.session_state.export_ready['excel']['data'],
                        file_name=st.session_state.export_ready['excel']['filename'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_excel_btn"
                    )
            else:
                st.info("📈 Excel export unavailable (Install openpyxl & xlsxwriter)")
        
        # New conversation
        if st.button("🆕 New Topic", use_container_width=True):
            st.session_state.messages = []
            st.session_state.research_assistant = ResearchGapAnalyzer()
            welcome = st.session_state.research_assistant.start_conversation()
            st.session_state.messages.append({
                'role': 'assistant',
                'content': welcome,
                'timestamp': datetime.now()
            })
            st.rerun()
        
        st.markdown("---")
        
        # Provider Status
        st.markdown("### AI Providers")
        
        providers = st.session_state.research_assistant.llm.get_available_providers()
        current = st.session_state.research_assistant.llm.get_current_provider()
        
        if providers:
            st.success(f"Active: {current}")
            for provider, available in providers.items():
                status = "🟢" if available else "🔴"
                st.text(f"{status} {provider.title()}")
        else:
            st.warning("No AI providers available")
            st.info("Basic search mode active")
        
        # Provider selection
        if providers:
            provider_list = list(providers.keys())
            provider_names = {
                'openai': 'OpenAI GPT-4',
                'anthropic': 'Anthropic Claude',
                'gemini': 'Google Gemini'
            }
            
            selected = st.selectbox(
                "Switch Provider",
                options=provider_list,
                format_func=lambda x: provider_names.get(x, x.title()),
                index=provider_list.index(current) if current in provider_list else 0,
                key="provider_selector"
            )
            
            if st.button("Activate", use_container_width=True):
                if st.session_state.research_assistant.llm.set_provider(selected):
                    st.success(f"✅ Switched to {provider_names.get(selected, selected)}")
                    time.sleep(0.5)  # Brief pause for user feedback
                    st.rerun()
                else:
                    st.error("Failed to switch provider")
        
        st.markdown("---")
        
        # Export
        st.markdown("### Export Data")
        
        if st.session_state.messages:
            export_data = {
                'conversation': [
                    {
                        'role': msg['role'],
                        'content': msg['content'],
                        'timestamp': msg['timestamp'].isoformat() if hasattr(msg['timestamp'], 'isoformat') else str(msg['timestamp'])
                    }
                    for msg in st.session_state.messages
                ],
                'research_context': {
                    'topic': st.session_state.research_assistant.current_topic,
                    'papers_found': len(st.session_state.research_assistant.recent_papers),
                    'gaps_identified': len(st.session_state.research_assistant.identified_gaps)
                },
                'export_timestamp': datetime.now().isoformat()
            }
            
            # JSON export
            json_data = json.dumps(export_data, indent=2)
            st.download_button(
                label="📥 Download JSON",
                data=json_data,
                file_name=f"research_gaps_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                use_container_width=True
            )
        
        st.markdown("---")
        
        # Stats
        st.markdown("### Session Stats")
        st.metric("Papers Found", len(st.session_state.research_assistant.recent_papers))
        st.metric("Gaps Identified", len(st.session_state.research_assistant.identified_gaps))
        st.metric("Messages", len(st.session_state.messages))
        
        # Help
        with st.expander("❓ Help"):
            st.markdown("""
            **How to use:**
            1. Type your research interest
            2. Click 'Find Papers' to search
            3. Click 'Analyze Gaps' for insights
            
            **Example topics:**
            • Machine learning fairness
            • Climate change adaptation
            • Mental health interventions
            • Renewable energy storage
            """)


# Main Application
def main():
    """Main application function with enhanced error handling"""
    try:
        initialize_session_state()
        display_header()
        
        # Create main layout
        chat_container = st.container()
        
        with chat_container:
            # Display messages
            for message in st.session_state.messages:
                try:
                    display_message(message)
                except Exception as e:
                    logger.error(f"Error displaying message: {e}")
                    st.error("Error displaying message")
            
            # Show thinking indicator if processing
            if st.session_state.thinking:
                display_thinking_indicator()
        
        # Enhanced input area with Enter key support and auto-clear
        user_input = st.chat_input(
            placeholder="Ask me about research topics, papers, or gaps... (e.g., 'machine learning in healthcare')",
            key="chat_input"
        )

        # Handle input submission (Enter key or send button)
        if user_input:
            send_message(user_input)
        
        # Enhanced input handling with Enter key support
        st.markdown("""
        <script>
        // Enhanced input handling
        setTimeout(() => {
            // Auto-focus on chat input
            const chatInput = document.querySelector('[data-testid="stChatInput"] textarea');
            if (chatInput) {
                chatInput.focus();

                // Ensure Enter key works (backup handling)
                chatInput.addEventListener('keydown', function(e) {
                    if (e.key === 'Enter' && !e.shiftKey) {
                        e.preventDefault();
                        // Trigger the send button if it exists
                        const sendButton = document.querySelector('[data-testid="stChatInput"] button');
                        if (sendButton) {
                            sendButton.click();
                        }
                    }
                });
            }

            // Fallback for text input if chat input not available
            const textInput = document.querySelector('input[type="text"]');
            if (textInput && !chatInput) {
                textInput.focus();
                textInput.addEventListener('keydown', function(e) {
                    if (e.key === 'Enter') {
                        e.preventDefault();
                        const sendButton = document.querySelector('button[kind="primary"]');
                        if (sendButton) {
                            sendButton.click();
                        }
                    }
                });
            }
        }, 100);
        </script>
        """, unsafe_allow_html=True)
        
        # Sidebar
        display_sidebar()
        
    except Exception as e:
        logger.error(f"Application error: {e}")
        st.error(f"An error occurred: {str(e)}")
        st.info("Try refreshing the page or checking your configuration.")

if __name__ == "__main__":
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Run the app
    main()